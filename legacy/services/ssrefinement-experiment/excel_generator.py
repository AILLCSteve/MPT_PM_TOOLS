"""
Excel generation module with three different visualization approaches.
Approach 1: openpyxl native charts
Approach 2: xlsxwriter charts
Approach 3: plotly charts embedded as images
"""

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Dict, List, Any
import io
import os


class ExcelDashboardGenerator:
    """Generate Excel dashboards with charts using different approaches."""

    def __init__(self, data_processor):
        """Initialize with a CIPPDataProcessor instance."""
        self.processor = data_processor
        self.tables = data_processor.get_all_tables()

    def generate_approach_1(self, output_path: str):
        """Approach 1: Pure openpyxl with native Excel charts."""
        wb = Workbook()

        # Create Dashboard_Data sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        ws_data = wb.create_sheet("Dashboard_Data", 0)
        self._write_all_tables(ws_data)

        # Create Dashboard sheet with charts
        ws_dash = wb.create_sheet("Dashboard", 1)
        self._create_openpyxl_charts(ws_dash, ws_data)

        wb.save(output_path)
        return output_path

    def generate_approach_2(self, output_path: str):
        """Approach 2: xlsxwriter with enhanced chart formatting."""
        try:
            import xlsxwriter
        except ImportError:
            raise ImportError("xlsxwriter not installed. Run: pip install xlsxwriter")

        wb = xlsxwriter.Workbook(output_path)

        # Create Dashboard_Data sheet
        ws_data = wb.add_worksheet("Dashboard_Data")
        self._write_all_tables_xlsxwriter(wb, ws_data)

        # Create Dashboard sheet with charts
        ws_dash = wb.add_worksheet("Dashboard")
        self._create_xlsxwriter_charts(wb, ws_dash, ws_data)

        wb.close()
        return output_path

    def generate_approach_3(self, output_path: str):
        """Approach 3: openpyxl data + plotly charts as images."""
        try:
            import plotly.graph_objects as go
            import plotly.express as px
            from PIL import Image as PILImage
        except ImportError:
            raise ImportError("plotly and/or Pillow not installed. Run: pip install plotly pillow kaleido")

        wb = Workbook()

        # Create Dashboard_Data sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        ws_data = wb.create_sheet("Dashboard_Data", 0)
        self._write_all_tables(ws_data)

        # Create Dashboard sheet with plotly charts as images
        ws_dash = wb.create_sheet("Dashboard", 1)
        self._create_plotly_charts(ws_dash, ws_data)

        wb.save(output_path)
        return output_path

    def _write_all_tables(self, ws):
        """Write all 5 tables to Dashboard_Data sheet using openpyxl."""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Table 1: Stage_Footage_Summary at A1
        self._write_table(ws, 1, 1, self.tables["stage_footage_summary"], header_fill, header_font)

        # Table 2: Stage_by_PipeSize at A10
        self._write_table(ws, 10, 1, self.tables["stage_by_pipe_size"], header_fill, header_font)

        # Table 3: Pipe_Size_Mix at A19
        self._write_table(ws, 19, 1, self.tables["pipe_size_mix"], header_fill, header_font)

        # Table 4: Length_Bins at A27
        self._write_table(ws, 27, 1, self.tables["length_bins"], header_fill, header_font)

        # Table 5: Easement_Traffic_Summary at A37
        self._write_table(ws, 37, 1, self.tables["easement_traffic_summary"], header_fill, header_font)

    def _write_table(self, ws, start_row, start_col, data, header_fill, header_font):
        """Write a single table to the worksheet."""
        if not data:
            return

        # Write headers
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers):
            cell = ws.cell(row=start_row, column=start_col + col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, header in enumerate(headers):
                value = row_data[header]
                cell = ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=value)

                # Format percentages
                if "Pct" in header or "%" in header:
                    cell.number_format = '0.00%'
                elif isinstance(value, (int, float)) and header != "Pipe Size":
                    cell.number_format = '#,##0.00'

    def _create_openpyxl_charts(self, ws_dash, ws_data):
        """Create charts using openpyxl native charts (Approach 1)."""
        # Add KPI cells at top
        ws_dash["A1"] = "CIPP Project Dashboard"
        ws_dash["A1"].font = Font(size=20, bold=True)

        ws_dash["A3"] = "Total Segments:"
        ws_dash["B3"] = sum(r["Segment_Count"] for r in self.tables["stage_footage_summary"])

        ws_dash["D3"] = "Total Footage:"
        ws_dash["E3"] = self.processor.total_footage

        # Chart 1: Overall Progress (100% stacked column)
        chart1 = BarChart()
        chart1.type = "col"
        chart1.grouping = "percentStacked"
        chart1.overlap = 100
        chart1.title = "Overall Progress by Stage"
        chart1.y_axis.title = "Percentage"
        chart1.x_axis.title = "Stage"

        # Data from Table 1 (A1:D7 in Dashboard_Data)
        data = Reference(ws_data, min_col=3, min_row=1, max_row=7, max_col=3)
        cats = Reference(ws_data, min_col=1, min_row=2, max_row=7)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)

        ws_dash.add_chart(chart1, "A5")

        # Chart 2: Progress by Pipe Size (stacked bar)
        chart2 = BarChart()
        chart2.type = "bar"
        chart2.grouping = "stacked"
        chart2.title = "Progress by Pipe Size"
        chart2.y_axis.title = "Pipe Size"
        chart2.x_axis.title = "Feet"

        # Find the range for Table 2
        table2_rows = len(self.tables["stage_by_pipe_size"]) + 1
        data = Reference(ws_data, min_col=2, min_row=10, max_row=10 + table2_rows - 1, max_col=7)
        cats = Reference(ws_data, min_col=1, min_row=11, max_row=10 + table2_rows - 1)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)

        ws_dash.add_chart(chart2, "J5")

        # Chart 3: Pipe Size Mix (clustered column)
        chart3 = BarChart()
        chart3.type = "col"
        chart3.grouping = "clustered"
        chart3.title = "Pipe Size Mix"
        chart3.y_axis.title = "Count / Footage"
        chart3.x_axis.title = "Pipe Size"

        table3_rows = len(self.tables["pipe_size_mix"]) + 1
        data = Reference(ws_data, min_col=2, min_row=19, max_row=19 + table3_rows - 1, max_col=3)
        cats = Reference(ws_data, min_col=1, min_row=20, max_row=19 + table3_rows - 1)
        chart3.add_data(data, titles_from_data=True)
        chart3.set_categories(cats)

        ws_dash.add_chart(chart3, "A25")

        # Chart 4: Length Distribution (column)
        chart4 = BarChart()
        chart4.type = "col"
        chart4.title = "Length Distribution"
        chart4.y_axis.title = "Total Feet"
        chart4.x_axis.title = "Length Bin"

        data = Reference(ws_data, min_col=5, min_row=27, max_row=32, max_col=5)
        cats = Reference(ws_data, min_col=1, min_row=28, max_row=32)
        chart4.add_data(data, titles_from_data=True)
        chart4.set_categories(cats)

        ws_dash.add_chart(chart4, "J25")

        # Chart 5: Easement & Traffic (clustered column)
        chart5 = BarChart()
        chart5.type = "col"
        chart5.grouping = "clustered"
        chart5.title = "Easement & Traffic Control"
        chart5.y_axis.title = "Total Feet"

        data = Reference(ws_data, min_col=4, min_row=37, max_row=41, max_col=4)
        cats = Reference(ws_data, min_col=1, min_row=38, max_row=41)
        chart5.add_data(data, titles_from_data=True)
        chart5.set_categories(cats)

        ws_dash.add_chart(chart5, "A45")

    def _write_all_tables_xlsxwriter(self, wb, ws):
        """Write all 5 tables using xlsxwriter."""
        # Define formats
        header_format = wb.add_format({
            'bold': True,
            'bg_color': '#4472C4',
            'font_color': 'white',
            'align': 'center'
        })

        number_format = wb.add_format({'num_format': '#,##0.00'})
        percent_format = wb.add_format({'num_format': '0.00%'})

        # Table 1: Stage_Footage_Summary at A1
        self._write_table_xlsxwriter(ws, 0, 0, self.tables["stage_footage_summary"],
                                      header_format, number_format, percent_format)

        # Table 2: Stage_by_PipeSize at A10
        self._write_table_xlsxwriter(ws, 9, 0, self.tables["stage_by_pipe_size"],
                                      header_format, number_format, percent_format)

        # Table 3: Pipe_Size_Mix at A19
        self._write_table_xlsxwriter(ws, 18, 0, self.tables["pipe_size_mix"],
                                      header_format, number_format, percent_format)

        # Table 4: Length_Bins at A27
        self._write_table_xlsxwriter(ws, 26, 0, self.tables["length_bins"],
                                      header_format, number_format, percent_format)

        # Table 5: Easement_Traffic_Summary at A37
        self._write_table_xlsxwriter(ws, 36, 0, self.tables["easement_traffic_summary"],
                                      header_format, number_format, percent_format)

    def _write_table_xlsxwriter(self, ws, start_row, start_col, data, header_fmt, num_fmt, pct_fmt):
        """Write a table using xlsxwriter."""
        if not data:
            return

        headers = list(data[0].keys())

        # Write headers
        for col_idx, header in enumerate(headers):
            ws.write(start_row, start_col + col_idx, header, header_fmt)

        # Write data
        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, header in enumerate(headers):
                value = row_data[header]

                if "Pct" in header or "%" in header:
                    ws.write(start_row + row_idx, start_col + col_idx, value, pct_fmt)
                elif isinstance(value, (int, float)) and header != "Pipe Size":
                    ws.write(start_row + row_idx, start_col + col_idx, value, num_fmt)
                else:
                    ws.write(start_row + row_idx, start_col + col_idx, value)

    def _create_xlsxwriter_charts(self, wb, ws_dash, ws_data):
        """Create charts using xlsxwriter (Approach 2)."""
        # Add KPI headers
        title_format = wb.add_format({'bold': True, 'font_size': 20})
        ws_dash.write('A1', 'CIPP Project Dashboard', title_format)

        ws_dash.write('A3', 'Total Segments:')
        ws_dash.write('B3', sum(r["Segment_Count"] for r in self.tables["stage_footage_summary"]))

        ws_dash.write('D3', 'Total Footage:')
        ws_dash.write('E3', self.processor.total_footage)

        # Chart 1: Overall Progress (100% stacked column)
        chart1 = wb.add_chart({'type': 'column', 'subtype': 'percent_stacked'})
        chart1.add_series({
            'name': '=Dashboard_Data!$C$1',
            'categories': '=Dashboard_Data!$A$2:$A$7',
            'values': '=Dashboard_Data!$C$2:$C$7',
            'data_labels': {'percentage': True}
        })
        chart1.set_title({'name': 'Overall Progress by Stage'})
        chart1.set_x_axis({'name': 'Stage'})
        chart1.set_y_axis({'name': 'Percentage'})
        chart1.set_style(11)
        ws_dash.insert_chart('A5', chart1, {'x_scale': 1.5, 'y_scale': 1.5})

        # Chart 2: Progress by Pipe Size (stacked bar)
        chart2 = wb.add_chart({'type': 'bar', 'subtype': 'stacked'})
        table2_rows = len(self.tables["stage_by_pipe_size"])

        for col_idx, stage in enumerate(self.processor.STAGES, start=2):
            col_letter = chr(65 + col_idx)  # B, C, D, etc.
            chart2.add_series({
                'name': f'=Dashboard_Data!${col_letter}$10',
                'categories': f'=Dashboard_Data!$A$11:$A${10 + table2_rows}',
                'values': f'=Dashboard_Data!${col_letter}$11:${col_letter}${10 + table2_rows}',
            })

        chart2.set_title({'name': 'Progress by Pipe Size'})
        chart2.set_x_axis({'name': 'Feet'})
        chart2.set_y_axis({'name': 'Pipe Size'})
        chart2.set_style(11)
        ws_dash.insert_chart('J5', chart2, {'x_scale': 1.5, 'y_scale': 1.5})

        # Chart 3: Pipe Size Mix (clustered column)
        chart3 = wb.add_chart({'type': 'column', 'subtype': 'clustered'})
        table3_rows = len(self.tables["pipe_size_mix"])

        chart3.add_series({
            'name': '=Dashboard_Data!$B$19',
            'categories': f'=Dashboard_Data!$A$20:$A${19 + table3_rows}',
            'values': f'=Dashboard_Data!$B$20:$B${19 + table3_rows}',
        })
        chart3.add_series({
            'name': '=Dashboard_Data!$C$19',
            'categories': f'=Dashboard_Data!$A$20:$A${19 + table3_rows}',
            'values': f'=Dashboard_Data!$C$20:$C${19 + table3_rows}',
        })

        chart3.set_title({'name': 'Pipe Size Mix'})
        chart3.set_x_axis({'name': 'Pipe Size'})
        chart3.set_y_axis({'name': 'Count / Footage'})
        chart3.set_style(11)
        ws_dash.insert_chart('A25', chart3, {'x_scale': 1.5, 'y_scale': 1.5})

        # Chart 4: Length Distribution
        chart4 = wb.add_chart({'type': 'column'})
        chart4.add_series({
            'name': '=Dashboard_Data!$E$27',
            'categories': '=Dashboard_Data!$A$28:$A$32',
            'values': '=Dashboard_Data!$E$28:$E$32',
            'data_labels': {'value': True}
        })
        chart4.set_title({'name': 'Length Distribution'})
        chart4.set_x_axis({'name': 'Length Bin'})
        chart4.set_y_axis({'name': 'Total Feet'})
        chart4.set_style(11)
        ws_dash.insert_chart('J25', chart4, {'x_scale': 1.5, 'y_scale': 1.5})

        # Chart 5: Easement & Traffic
        chart5 = wb.add_chart({'type': 'column', 'subtype': 'clustered'})
        chart5.add_series({
            'name': '=Dashboard_Data!$D$37',
            'categories': '=Dashboard_Data!$A$38:$A$41',
            'values': '=Dashboard_Data!$D$38:$D$41',
            'data_labels': {'value': True}
        })
        chart5.set_title({'name': 'Easement & Traffic Control'})
        chart5.set_x_axis({'name': 'Category'})
        chart5.set_y_axis({'name': 'Total Feet'})
        chart5.set_style(11)
        ws_dash.insert_chart('A45', chart5, {'x_scale': 1.5, 'y_scale': 1.5})

    def _create_plotly_charts(self, ws_dash, ws_data):
        """Create charts using plotly and embed as images (Approach 3)."""
        import plotly.graph_objects as go
        import plotly.express as px

        # Add KPI cells
        ws_dash["A1"] = "CIPP Project Dashboard"
        ws_dash["A1"].font = Font(size=20, bold=True)

        ws_dash["A3"] = "Total Segments:"
        ws_dash["B3"] = sum(r["Segment_Count"] for r in self.tables["stage_footage_summary"])

        ws_dash["D3"] = "Total Footage:"
        ws_dash["E3"] = self.processor.total_footage

        # Chart 1: Overall Progress (100% stacked)
        table1 = self.tables["stage_footage_summary"]
        fig1 = go.Figure(data=[
            go.Bar(
                x=[r["Stage"] for r in table1],
                y=[r["Total_Feet"] for r in table1],
                text=[f"{r['Pct_of_Total_Feet']*100:.1f}%" for r in table1],
                textposition='auto',
                marker_color='#4472C4'
            )
        ])
        fig1.update_layout(
            title="Overall Progress by Stage",
            xaxis_title="Stage",
            yaxis_title="Total Feet",
            template="plotly_white",
            height=400,
            width=600
        )

        # Save and embed chart 1
        img_bytes1 = fig1.to_image(format="png")
        img1 = XLImage(io.BytesIO(img_bytes1))
        ws_dash.add_image(img1, "A5")

        # Chart 2: Progress by Pipe Size (stacked)
        table2 = self.tables["stage_by_pipe_size"]
        fig2 = go.Figure()

        for stage in self.processor.STAGES:
            fig2.add_trace(go.Bar(
                name=stage,
                y=[str(r["Pipe Size"]) for r in table2],
                x=[r[stage] for r in table2],
                orientation='h'
            ))

        fig2.update_layout(
            barmode='stack',
            title="Progress by Pipe Size",
            xaxis_title="Feet",
            yaxis_title="Pipe Size",
            template="plotly_white",
            height=400,
            width=700
        )

        img_bytes2 = fig2.to_image(format="png")
        img2 = XLImage(io.BytesIO(img_bytes2))
        ws_dash.add_image(img2, "K5")

        # Chart 3: Pipe Size Mix
        table3 = self.tables["pipe_size_mix"]
        fig3 = go.Figure()

        fig3.add_trace(go.Bar(
            name='Segment Count',
            x=[str(r["Pipe Size"]) for r in table3],
            y=[r["Segment_Count"] for r in table3]
        ))

        fig3.add_trace(go.Bar(
            name='Total Feet',
            x=[str(r["Pipe Size"]) for r in table3],
            y=[r["Total_Feet"] for r in table3]
        ))

        fig3.update_layout(
            barmode='group',
            title="Pipe Size Mix",
            xaxis_title="Pipe Size",
            yaxis_title="Count / Footage",
            template="plotly_white",
            height=400,
            width=600
        )

        img_bytes3 = fig3.to_image(format="png")
        img3 = XLImage(io.BytesIO(img_bytes3))
        ws_dash.add_image(img3, "A30")

        # Chart 4: Length Distribution
        table4 = self.tables["length_bins"]
        fig4 = go.Figure(data=[
            go.Bar(
                x=[r["Length_Bin_Label"] for r in table4],
                y=[r["Total_Feet"] for r in table4],
                text=[r["Total_Feet"] for r in table4],
                textposition='auto',
                marker_color='#70AD47'
            )
        ])

        fig4.update_layout(
            title="Length Distribution",
            xaxis_title="Length Bin",
            yaxis_title="Total Feet",
            template="plotly_white",
            height=400,
            width=600
        )

        img_bytes4 = fig4.to_image(format="png")
        img4 = XLImage(io.BytesIO(img_bytes4))
        ws_dash.add_image(img4, "K30")

        # Chart 5: Easement & Traffic
        table5 = self.tables["easement_traffic_summary"]
        fig5 = go.Figure(data=[
            go.Bar(
                x=[f"{r['Category']} - {r['Flag']}" for r in table5],
                y=[r["Total_Feet"] for r in table5],
                text=[f"{r['Pct_of_Total_Feet']*100:.1f}%" for r in table5],
                textposition='auto',
                marker_color=['#E7E6E6' if r['Flag'] == 'No' else '#FFC000' for r in table5]
            )
        ])

        fig5.update_layout(
            title="Easement & Traffic Control",
            xaxis_title="Category",
            yaxis_title="Total Feet",
            template="plotly_white",
            height=400,
            width=700
        )

        img_bytes5 = fig5.to_image(format="png")
        img5 = XLImage(io.BytesIO(img_bytes5))
        ws_dash.add_image(img5, "A55")
