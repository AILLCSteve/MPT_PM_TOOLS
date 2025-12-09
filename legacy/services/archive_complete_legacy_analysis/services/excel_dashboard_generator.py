"""
Excel Dashboard Generator with openpyxl
Creates professional Excel dashboards with embedded charts and visualizations.

Uses openpyxl to generate native Excel charts (Pie, Bar, Line, Scatter)
from HOTDOG AI analysis results.
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
from openpyxl import Workbook
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelDashboardGenerator:
    """
    Generates professional Excel dashboards with embedded charts.

    Creates multiple worksheets:
    1. Executive Dashboard - Overview with charts
    2. Detailed Results - Question-answer table
    3. Section Analysis - Section-wise breakdown
    4. Confidence Analysis - Confidence distribution
    5. Footnotes - All page citations
    """

    def __init__(self):
        """Initialize the dashboard generator."""
        self.wb = None
        self.primary_color = "5B7FCC"  # MPT Blue
        self.secondary_color = "1E3A8A"  # Dark blue
        self.success_color = "22C55E"  # Green
        self.warning_color = "F59E0B"  # Amber
        self.danger_color = "EF4444"  # Red

    def generate_dashboard(
        self,
        analysis_result: Dict[str, Any],
        output_path: str,
        document_name: str = "Document Analysis"
    ) -> str:
        """
        Generate complete Excel dashboard with charts.

        Args:
            analysis_result: Analysis result from HOTDOG (browser format)
            output_path: Path to save Excel file
            document_name: Name of analyzed document

        Returns:
            Path to generated Excel file
        """
        logger.info(f"📊 Generating Excel dashboard: {output_path}")

        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet

        # Extract data
        sections = analysis_result.get('sections', [])
        total_questions = analysis_result.get('total_questions', 0)
        questions_answered = analysis_result.get('questions_answered', 0)
        questions_unanswered = total_questions - questions_answered

        # Create worksheets
        logger.info("Creating Executive Dashboard...")
        self._create_executive_dashboard(sections, total_questions, questions_answered, document_name)

        logger.info("Creating Detailed Results...")
        self._create_detailed_results(sections)

        logger.info("Creating Section Analysis...")
        self._create_section_analysis(sections)

        logger.info("Creating Confidence Analysis...")
        self._create_confidence_analysis(sections)

        logger.info("Creating Footnotes...")
        self._create_footnotes(analysis_result.get('footnotes', []))

        # Save workbook
        self.wb.save(output_path)
        logger.info(f"✅ Excel dashboard saved: {output_path}")

        return output_path

    def _create_executive_dashboard(
        self,
        sections: List[Dict],
        total_questions: int,
        questions_answered: int,
        document_name: str
    ):
        """Create executive dashboard with overview charts."""
        ws = self.wb.create_sheet("Executive Dashboard")

        # Header
        ws['A1'] = "CIPP ANALYSIS DASHBOARD"
        ws['A1'].font = Font(size=20, bold=True, color=self.primary_color)
        ws.merge_cells('A1:F1')

        ws['A2'] = f"Document: {document_name}"
        ws['A2'].font = Font(size=12, italic=True)
        ws.merge_cells('A2:F2')

        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A3'].font = Font(size=10, color="666666")
        ws.merge_cells('A3:F3')

        # Completion Rate Data
        ws['A5'] = "COMPLETION OVERVIEW"
        ws['A5'].font = Font(size=14, bold=True)

        ws['A6'] = "Category"
        ws['B6'] = "Count"
        self._style_header_row(ws, 6, 2)

        questions_unanswered = total_questions - questions_answered
        ws['A7'] = "Answered"
        ws['B7'] = questions_answered
        ws['A8'] = "Unanswered"
        ws['B8'] = questions_unanswered

        # Create Pie Chart for completion
        pie = PieChart()
        pie.title = "Analysis Completion Rate"
        pie.style = 10
        pie.height = 10  # Height in cm
        pie.width = 15   # Width in cm

        labels = Reference(ws, min_col=1, min_row=7, max_row=8)
        data = Reference(ws, min_col=2, min_row=6, max_row=8)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)

        # Add data labels
        from openpyxl.chart.label import DataLabelList
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCatName = True

        ws.add_chart(pie, "D5")

        # Section Performance Data
        ws['A20'] = "SECTION PERFORMANCE"
        ws['A20'].font = Font(size=14, bold=True)

        ws['A21'] = "Section"
        ws['B21'] = "Answered"
        ws['C21'] = "Total"
        ws['D21'] = "Completion %"
        self._style_header_row(ws, 21, 4)

        row = 22
        for section in sections:
            section_name = section.get('name', 'Unknown')
            answered = len([q for q in section.get('questions', []) if q.get('answer') and q['answer'] != 'Not yet found.'])
            total = len(section.get('questions', []))
            completion = (answered / total * 100) if total > 0 else 0

            ws[f'A{row}'] = section_name
            ws[f'B{row}'] = answered
            ws[f'C{row}'] = total
            ws[f'D{row}'] = f"{completion:.1f}%"
            row += 1

        # Create Bar Chart for section performance
        bar = BarChart()
        bar.title = "Section Completion Rates"
        bar.style = 10
        bar.height = 10
        bar.width = 15
        bar.x_axis.title = "Sections"
        bar.y_axis.title = "Questions Answered"

        categories = Reference(ws, min_col=1, min_row=22, max_row=row-1)
        data = Reference(ws, min_col=2, min_row=21, max_row=row-1)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(categories)

        ws.add_chart(bar, "F20")

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15

    def _create_detailed_results(self, sections: List[Dict]):
        """Create detailed results table."""
        ws = self.wb.create_sheet("Detailed Results")

        # Header
        headers = ["Section", "Q#", "Question", "Answer", "Confidence", "Page Citations"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.primary_color, end_color=self.primary_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        row = 2
        for section in sections:
            section_name = section.get('name', 'Unknown')
            questions = section.get('questions', [])

            for q in questions:
                ws.cell(row, 1, section_name)
                ws.cell(row, 2, q.get('id', ''))
                ws.cell(row, 3, q.get('text', ''))

                answer = q.get('answer', 'Not yet found.')
                ws.cell(row, 4, answer)

                confidence = q.get('confidence', 0.0)
                ws.cell(row, 5, f"{confidence:.0%}" if confidence > 0 else "N/A")

                # Color code confidence
                conf_cell = ws.cell(row, 5)
                if confidence >= 0.7:
                    conf_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif confidence >= 0.4:
                    conf_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif confidence > 0:
                    conf_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                pages = q.get('pages', [])
                ws.cell(row, 6, ', '.join(map(str, pages)) if pages else "")

                row += 1

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20

        # Freeze header row
        ws.freeze_panes = 'A2'

    def _create_section_analysis(self, sections: List[Dict]):
        """Create section-wise analysis with charts."""
        ws = self.wb.create_sheet("Section Analysis")

        # Header
        ws['A1'] = "SECTION-WISE ANALYSIS"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')

        # Data headers
        ws['A3'] = "Section"
        ws['B3'] = "Total Questions"
        ws['C3'] = "Answered"
        ws['D3'] = "Unanswered"
        ws['E3'] = "Completion %"
        self._style_header_row(ws, 3, 5)

        # Data
        row = 4
        for section in sections:
            section_name = section.get('name', 'Unknown')
            questions = section.get('questions', [])
            total = len(questions)
            answered = len([q for q in questions if q.get('answer') and q['answer'] != 'Not yet found.'])
            unanswered = total - answered
            completion = (answered / total * 100) if total > 0 else 0

            ws[f'A{row}'] = section_name
            ws[f'B{row}'] = total
            ws[f'C{row}'] = answered
            ws[f'D{row}'] = unanswered
            ws[f'E{row}'] = f"{completion:.1f}%"

            # Color code completion
            comp_cell = ws[f'E{row}']
            if completion >= 90:
                comp_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif completion >= 70:
                comp_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                comp_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15

    def _create_confidence_analysis(self, sections: List[Dict]):
        """Create confidence distribution analysis."""
        ws = self.wb.create_sheet("Confidence Analysis")

        # Collect confidence data
        high_conf = 0
        medium_conf = 0
        low_conf = 0
        no_answer = 0

        for section in sections:
            for q in section.get('questions', []):
                confidence = q.get('confidence', 0.0)
                if confidence >= 0.7:
                    high_conf += 1
                elif confidence >= 0.4:
                    medium_conf += 1
                elif confidence > 0:
                    low_conf += 1
                else:
                    no_answer += 1

        # Header
        ws['A1'] = "CONFIDENCE DISTRIBUTION"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')

        # Data
        ws['A3'] = "Confidence Level"
        ws['B3'] = "Count"
        ws['C3'] = "Threshold"
        self._style_header_row(ws, 3, 3)

        ws['A4'] = "High Confidence"
        ws['B4'] = high_conf
        ws['C4'] = "≥ 70%"

        ws['A5'] = "Medium Confidence"
        ws['B5'] = medium_conf
        ws['C5'] = "40% - 70%"

        ws['A6'] = "Low Confidence"
        ws['B6'] = low_conf
        ws['C6'] = "< 40%"

        ws['A7'] = "No Answer"
        ws['B7'] = no_answer
        ws['C7'] = "0%"

        # Create Bar Chart
        bar = BarChart()
        bar.title = "Confidence Distribution"
        bar.style = 10
        bar.height = 12
        bar.width = 18

        categories = Reference(ws, min_col=1, min_row=4, max_row=7)
        data = Reference(ws, min_col=2, min_row=3, max_row=7)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(categories)

        ws.add_chart(bar, "E3")

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15

    def _create_footnotes(self, footnotes: List[str]):
        """Create footnotes sheet."""
        ws = self.wb.create_sheet("Footnotes")

        # Header
        ws['A1'] = "PAGE CITATIONS & FOOTNOTES"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')

        ws['A3'] = "#"
        ws['B3'] = "Citation"
        self._style_header_row(ws, 3, 2)

        # Data
        for i, footnote in enumerate(footnotes, 1):
            ws[f'A{i+3}'] = i
            ws[f'B{i+3}'] = footnote

        # Column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 80

    def _style_header_row(self, ws, row: int, cols: int):
        """Apply consistent header styling."""
        for col in range(1, cols + 1):
            cell = ws.cell(row, col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.primary_color, end_color=self.primary_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")


def generate_excel_dashboard(
    analysis_result: Dict[str, Any],
    output_path: str,
    document_name: str = "Document Analysis"
) -> str:
    """
    Convenience function to generate Excel dashboard.

    Args:
        analysis_result: Analysis result from HOTDOG (browser format)
        output_path: Path to save Excel file
        document_name: Name of analyzed document

    Returns:
        Path to generated Excel file
    """
    generator = ExcelDashboardGenerator()
    return generator.generate_dashboard(analysis_result, output_path, document_name)
