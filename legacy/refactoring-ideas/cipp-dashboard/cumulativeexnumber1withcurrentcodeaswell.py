def _compute_stage(self, segment: Dict[str, Any]) -> str:
    if segment["final_post_tv_date"] is not None:
        return "Post TV Complete"
    elif segment["lining_date"] is not None:
        return "Lined"
    elif segment["ready_to_line"]:
        return "Ready to Line"
    elif segment["prep_complete"]:
        return "Prep Complete"
    else:
        return "Not Started"
``` :contentReference[oaicite:0]{index=0}  

Then `get_stage_footage_summary` just **groups by that single stage** and computes percentages of total feet. :contentReference[oaicite:1]{index=1}  

Dash and Excel use that summary as the lifecycle view (e.g., `LIFECYCLE_STAGES = ['Prep Complete', 'Ready to Line', 'Lined', 'Post TV Complete']` in `dash_app.py`). :contentReference[oaicite:2]{index=2}  

So when a segment moves from “Ready to Line” to “Lined”, its footage is **removed** from Ready to Line and **added** to Lined. The stacked bar and percentages always show the *current* stage only – which is perfect for “where is everything right now?”, but terrible for “how much prep have we completed over the life of the job?”.

We’re going to introduce a **cumulative milestone model** on top of this so:

- A segment that is “Post TV Complete” still **counts as completed** for:
  - Prep Complete
  - Ready to Line
  - Wet Out
  - Lined
  - Post TV Complete  

…instead of “forgetting” earlier milestones when it advances.

---

## 2. Refactored example #1 – lifecycle-aware data processor (with Wet Out, cumulative milestones, potential issues, and lifecycle segment tables)

**File name suggestion:**  
`services/cipp_dashboard/data_processor_lifecycle_v2.py`

This is a **drop-in alternative** to feed to Claude. It keeps the structure of your existing `CIPPDataProcessor` but:

- Adds **`Wet Out`** as a milestone (pulled from a date column)
- Keeps the old exclusive `stage` but adds **cumulative milestone coverage**
- Adds a **Potential Issues** table (Prep Complete == True, Ready to Line == False)
- Adds scaffold for **per-lifecycle tables** (Ready to Line, Not Started, Lined, Post TV Complete)

```python
"""
Lifecycle-aware variant of CIPPDataProcessor.

Key changes:
- Adds 'Wet Out' milestone from a date column.
- Introduces cumulative lifecycle milestones (Prep → Ready → Wet Out → Lined → Post TV).
- Adds 'Potential Issues' table: Prep Complete = True AND Ready to Line = False.
- Adds lifecycle segment tables (e.g. all Ready to Line segments).
"""

from collections import defaultdict
from typing import List, Dict, Any, Optional

from openpyxl import load_workbook


class CIPPDataProcessorLifecycleV2:
    """
    Refined data processor that keeps the original 'stage' concept
    (current stage) but adds lifecycle-aware cumulative milestones.
    """

    # Note: Prep Complete remains in STAGES for backward compatibility, but we
    # will EXCLUDE it from the lifecycle stacked bar if desired.
    STAGES = [
        "Not Started",
        "Prep Complete",
        "Ready to Line",
        "Wet Out",
        "Lined",
        "Post TV Complete",
    ]

    # Lifecycle milestones in chronological order
    LIFECYCLE_MILESTONES = [
        "Prep Complete",
        "Ready to Line",
        "Wet Out",
        "Lined",
        "Post TV Complete",
    ]

    LENGTH_BINS = [
        {"label": "0–50", "min": 0, "max": 50},
        {"label": "51–150", "min": 51, "max": 150},
        {"label": "151–250", "min": 151, "max": 250},
        {"label": "251–350", "min": 251, "max": 350},
        {"label": "351+", "min": 351, "max": None},
    ]

    def __init__(self, file_path: str, sheet_name: str = "WEST DES MOINES, IA Shot Schedu"):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.segments: List[Dict[str, Any]] = []
        self.total_footage: float = 0.0

    # --------------------------------------------------------------------- #
    # DATA LOAD
    # --------------------------------------------------------------------- #

    def load_data(self) -> List[Dict[str, Any]]:
        """Load and validate data from Excel file (adds Wet Out)."""
        wb = load_workbook(self.file_path, data_only=True)

        # Try to find the sheet
        if self.sheet_name not in wb.sheetnames:
            for sheet in wb.sheetnames:
                if "MOINES" in sheet.upper() or "SHOT" in sheet.upper():
                    self.sheet_name = sheet
                    break
            else:
                raise ValueError(f"Sheet '{self.sheet_name}' not found. Available: {wb.sheetnames}")

        ws = wb[self.sheet_name]

        # Build header map from row 1
        headers: Dict[str, int] = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx

        # Try a couple of possible Wet Out header names
        wet_out_header = (
            headers.get("Wet Out")
            or headers.get("Wet Out Date")
            or headers.get("Wet-Out")
        )

        segments: List[Dict[str, Any]] = []

        for row_idx in range(2, ws.max_row + 1):
            row = ws[row_idx]

            video_id = self._get_cell_value(row, headers.get("VIDEO ID"))
            map_length = self._get_cell_value(row, headers.get("Map Length"))

            if not self._is_valid_segment(video_id, map_length):
                continue

            segment: Dict[str, Any] = {
                "video_id": video_id,
                "line_segment": self._get_cell_value(row, headers.get("Line Segment")),
                "pipe_size": self._get_cell_value(row, headers.get("Pipe Size")),
                "map_length": float(map_length),
                "prep_complete": self._is_truthy(
                    self._get_cell_value(row, headers.get("Prep Complete"))
                ),
                "ready_to_line": self._is_truthy(
                    self._get_cell_value(
                        row,
                        headers.get(
                            "Ready to Line - Certified by Prep Crew Lead"
                        ),
                    )
                ),
                "lining_date": self._get_cell_value(row, headers.get("Lining Date")),
                "final_post_tv_date": self._get_cell_value(
                    row, headers.get("Final Post TV Date")
                ),
                "grout_state_date": self._get_cell_value(
                    row, headers.get("Grout State Date")
                ),
                "easement": self._is_truthy(
                    self._get_cell_value(row, headers.get("Easement"))
                ),
                "traffic_control": self._is_truthy(
                    self._get_cell_value(row, headers.get("Traffic Control"))
                ),
            }

            # New: Wet Out date
            segment["wet_out_date"] = (
                self._get_cell_value(row, wet_out_header) if wet_out_header else None
            )

            # Keep the notion of "current stage" for compatibility
            segment["stage"] = self._compute_stage(segment)
            segments.append(segment)

        self.segments = segments
        self.total_footage = sum(s["map_length"] for s in segments)
        return segments

    # --------------------------------------------------------------------- #
    # HELPER METHODS
    # --------------------------------------------------------------------- #

    def _get_cell_value(self, row, col_idx: Optional[int]):
        if col_idx is None or col_idx < 1:
            return None
        try:
            return row[col_idx - 1].value
        except (IndexError, AttributeError):
            return None

    def _is_valid_segment(self, video_id, map_length) -> bool:
        try:
            video_id_num = float(video_id) if video_id is not None else 0
            map_length_num = float(map_length) if map_length is not None else 0
            return video_id_num >= 1 and map_length_num > 0
        except (ValueError, TypeError):
            return False

    def _is_truthy(self, value) -> bool:
        if value is None:
            return False
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return value != 0
        if isinstance(value, str):
            return value.upper() in ["TRUE", "YES", "Y", "1"]
        return bool(value)

    # --------------------------------------------------------------------- #
    # STAGE + MILESTONES
    # --------------------------------------------------------------------- #

    def _compute_stage(self, segment: Dict[str, Any]) -> str:
        """
        Current stage priority:
        Post TV > Lined > Wet Out > Ready to Line > Prep Complete > Not Started
        """
        if segment.get("final_post_tv_date") is not None:
            return "Post TV Complete"
        if segment.get("lining_date") is not None:
            return "Lined"
        if segment.get("wet_out_date") is not None:
            return "Wet Out"
        if segment.get("ready_to_line"):
            return "Ready to Line"
        if segment.get("prep_complete"):
            return "Prep Complete"
        return "Not Started"

    # --- Original, exclusive stage summary (still useful) ---------------- #

    def get_stage_footage_summary(self) -> List[Dict[str, Any]]:
        stage_data = defaultdict(lambda: {"count": 0, "footage": 0.0})
        for segment in self.segments:
            stage = segment["stage"]
            stage_data[stage]["count"] += 1
            stage_data[stage]["footage"] += segment["map_length"]

        result: List[Dict[str, Any]] = []
        for stage in self.STAGES:
            data = stage_data[stage]
            result.append(
                {
                    "Stage": stage,
                    "Segment_Count": data["count"],
                    "Total_Feet": round(data["footage"], 2),
                    "Pct_of_Total_Feet": round(
                        data["footage"] / self.total_footage, 4
                    )
                    if self.total_footage > 0
                    else 0,
                }
            )
        return result

    # --- NEW: cumulative lifecycle milestones ---------------------------- #

    def get_lifecycle_milestones(self) -> List[Dict[str, Any]]:
        """
        For each milestone M_k, compute coverage as:
        - Cumulative_Feet: sum of footage for segments that have reached M_k or later.
        - Cumulative_Segment_Count: count of those segments.
        Then also compute "delta" width so stacked bars still sum to 100%.
        """
        if not self.segments or self.total_footage <= 0:
            return []

        # Assign numeric index to each milestone
        milestone_index = {name: i for i, name in enumerate(self.LIFECYCLE_MILESTONES)}

        # For each segment, find the highest milestone achieved
        max_index_for_segment: List[int] = []
        for seg in self.segments:
            highest = -1
            if seg.get("prep_complete"):
                highest = max(highest, milestone_index["Prep Complete"])
            if seg.get("ready_to_line"):
                highest = max(highest, milestone_index["Ready to Line"])
            if seg.get("wet_out_date") is not None:
                highest = max(highest, milestone_index["Wet Out"])
            if seg.get("lining_date") is not None:
                highest = max(highest, milestone_index["Lined"])
            if seg.get("final_post_tv_date") is not None:
                highest = max(highest, milestone_index["Post TV Complete"])

            max_index_for_segment.append(highest)

        # Cumulative coverage by milestone (feet + count)
        cumulative_feet = [0.0] * len(self.LIFECYCLE_MILESTONES)
        cumulative_count = [0] * len(self.LIFECYCLE_MILESTONES)

        for seg, max_idx in zip(self.segments, max_index_for_segment):
            if max_idx < 0:
                # Segment has not reached any milestone
                continue
            length = seg["map_length"]
            for i in range(max_idx + 1):
                cumulative_feet[i] += length
                cumulative_count[i] += 1

        # Convert to rows and compute deltas so we can still draw stacked bars
        rows: List[Dict[str, Any]] = []
        prev_pct = 0.0
        for i, name in enumerate(self.LIFECYCLE_MILESTONES):
            feet = cumulative_feet[i]
            pct_cum = feet / self.total_footage if self.total_footage > 0 else 0.0
            delta_pct = max(pct_cum - prev_pct, 0.0)
            prev_pct = pct_cum

            rows.append(
                {
                    "Stage": name,
                    # cumulative coverage (what % of the project has EVER reached this milestone)
                    "Cumulative_Segment_Count": cumulative_count[i],
                    "Cumulative_Feet": round(feet, 2),
                    "Cumulative_Pct_of_Total_Feet": round(pct_cum, 4),
                    # incremental width used for stacked bars so total stays at 100%
                    "Delta_Pct_of_Total_Feet": round(delta_pct, 4),
                }
            )

        return rows

    # --------------------------------------------------------------------- #
    # POTENTIAL ISSUES TABLE
    # --------------------------------------------------------------------- #

    def get_potential_issues(self) -> List[Dict[str, Any]]:
        """
        Segments where prep is complete but the segment is not yet Ready to Line.
        These are likely bottlenecks or QA/coordination issues.
        """
        issues: List[Dict[str, Any]] = []

        for s in self.segments:
            if s.get("prep_complete") and not s.get("ready_to_line"):
                issues.append(
                    {
                        "Video_ID": s["video_id"],
                        "Line_Segment": s["line_segment"],
                        "Pipe_Size": s["pipe_size"],
                        "Map_Length_ft": s["map_length"],
                        "Prep_Complete": s["prep_complete"],
                        "Ready_to_Line": s["ready_to_line"],
                        "Wet_Out_Date": s.get("wet_out_date"),
                        "Lining_Date": s.get("lining_date"),
                        "Final_Post_TV_Date": s.get("final_post_tv_date"),
                        "Easement": s.get("easement"),
                        "Traffic_Control": s.get("traffic_control"),
                    }
                )

        return issues

    # --------------------------------------------------------------------- #
    # LIFECYCLE SEGMENT TABLES (SCaffold)
    # --------------------------------------------------------------------- #

    def _build_segment_table(self, filter_func, label: str) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        for s in self.segments:
            if not filter_func(s):
                continue
            rows.append(
                {
                    "Lifecycle_Group": label,
                    "Video_ID": s["video_id"],
                    "Line_Segment": s["line_segment"],
                    "Pipe_Size": s["pipe_size"],
                    "Map_Length_ft": s["map_length"],
                    "Stage": s["stage"],
                    "Prep_Complete": s["prep_complete"],
                    "Ready_to_Line": s["ready_to_line"],
                    "Wet_Out_Date": s.get("wet_out_date"),
                    "Lining_Date": s.get("lining_date"),
                    "Final_Post_TV_Date": s.get("final_post_tv_date"),
                    "Easement": s.get("easement"),
                    "Traffic_Control": s.get("traffic_control"),
                }
            )
        return rows

    def get_lifecycle_segment_tables(self) -> Dict[str, List[Dict[str, Any]]]:
        """
        Return grouped segment tables for key lifecycle views, e.g.:
        - Not Yet Started
        - Ready to Line
        - Lined
        - CCTV Post Complete
        """
        tables: Dict[str, List[Dict[str, Any]]] = {}

        tables["Not Started"] = self._build_segment_table(
            lambda s: s["stage"] == "Not Started", "Not Started"
        )
        tables["Ready to Line"] = self._build_segment_table(
            lambda s: s.get("ready_to_line"), "Ready to Line"
        )
        tables["Lined"] = self._build_segment_table(
            lambda s: s.get("lining_date") is not None, "Lined"
        )
        tables["Post TV Complete"] = self._build_segment_table(
            lambda s: s.get("final_post_tv_date") is not None, "Post TV Complete"
        )

        # You can add "Wet Out" or other specialized views here as needed.
        tables["Wet Out"] = self._build_segment_table(
            lambda s: s.get("wet_out_date") is not None, "Wet Out"
        )

        return tables

    # --------------------------------------------------------------------- #
    # EXISTING TABLES + NEW ONES
    # --------------------------------------------------------------------- #

    def get_stage_by_pipe_size(self) -> List[Dict[str, Any]]:
        pipe_sizes = sorted(
            set(s["pipe_size"] for s in self.segments if s["pipe_size"] is not None)
        )

        result: List[Dict[str, Any]] = []
        for pipe_size in pipe_sizes:
            row: Dict[str, Any] = {"Pipe Size": pipe_size}
            total = 0.0

            for stage in self.STAGES:
                footage = sum(
                    s["map_length"]
                    for s in self.segments
                    if s["pipe_size"] == pipe_size and s["stage"] == stage
                )
                row[stage] = round(footage, 2)
                total += footage

            row["Total_Feet"] = round(total, 2)
            result.append(row)

        return result

    def get_pipe_size_mix(self) -> List[Dict[str, Any]]:
        pipe_sizes = sorted(
            set(s["pipe_size"] for s in self.segments if s["pipe_size"] is not None)
        )

        result: List[Dict[str, Any]] = []
        for pipe_size in pipe_sizes:
            segs = [s for s in self.segments if s["pipe_size"] == pipe_size]
            count = len(segs)
            footage = sum(s["map_length"] for s in segs)

            result.append(
                {
                    "Pipe Size": pipe_size,
                    "Segment_Count": count,
                    "Total_Feet": round(footage, 2),
                    "Avg_Length_ft": round(footage / count, 2) if count > 0 else 0,
                    "Pct_of_Total_Feet": round(
                        footage / self.total_footage, 4
                    )
                    if self.total_footage > 0
                    else 0,
                }
            )

        return result

    def get_length_bins(self) -> List[Dict[str, Any]]:
        result: List[Dict[str, Any]] = []

        for bin_def in self.LENGTH_BINS:
            segments_in_bin: List[Dict[str, Any]] = []
            for segment in self.segments:
                length = segment["map_length"]
                min_len = bin_def["min"]
                max_len = bin_def["max"]

                if max_len is None:
                    if length >= min_len:
                        segments_in_bin.append(segment)
                else:
                    if min_len <= length <= max_len:
                        segments_in_bin.append(segment)

            count = len(segments_in_bin)
            footage = sum(s["map_length"] for s in segments_in_bin)

            result.append(
                {
                    "Length_Bin_Label": bin_def["label"],
                    "Min_Length_ft": bin_def["min"],
                    "Max_Length_ft": bin_def["max"] if bin_def["max"] is not None else "",
                    "Segment_Count": count,
                    "Total_Feet": round(footage, 2),
                    "Pct_of_Total_Feet": round(
                        footage / self.total_footage, 4
                    )
                    if self.total_footage > 0
                    else 0,
                }
            )

        return result

    def get_easement_traffic_summary(self) -> List[Dict[str, Any]]:
        result: List[Dict[str, Any]] = []

        easement_yes = [s for s in self.segments if s["easement"]]
        easement_no = [s for s in self.segments if not s["easement"]]
        traffic_yes = [s for s in self.segments if s["traffic_control"]]
        traffic_no = [s for s in self.segments if not s["traffic_control"]]

        def build_row(category: str, flag: str, segs: List[Dict[str, Any]]):
            footage = sum(s["map_length"] for s in segs)
            return {
                "Category": category,
                "Flag": flag,
                "Segment_Count": len(segs),
                "Total_Feet": round(footage, 2),
                "Pct_of_Total_Feet": round(
                    footage / self.total_footage, 4
                )
                if self.total_footage > 0
                else 0,
            }

        result.append(build_row("Easement", "Yes", easement_yes))
        result.append(build_row("Easement", "No", easement_no))
        result.append(build_row("Traffic Control", "Yes", traffic_yes))
        result.append(build_row("Traffic Control", "No", traffic_no))

        return result

    def get_all_tables(self) -> Dict[str, Any]:
        """
        All tables for Dash/Excel.
        - Keeps original ones for backward compatibility.
        - Adds lifecycle_milestones, potential_issues, lifecycle_segment_tables.
        """
        return {
            "stage_footage_summary": self.get_stage_footage_summary(),
            "lifecycle_milestones": self.get_lifecycle_milestones(),
            "stage_by_pipe_size": self.get_stage_by_pipe_size(),
            "pipe_size_mix": self.get_pipe_size_mix(),
            "length_bins": self.get_length_bins(),
            "easement_traffic_summary": self.get_easement_traffic_summary(),
            "potential_issues": self.get_potential_issues(),
            "lifecycle_segment_tables": self.get_lifecycle_segment_tables(),
        }
