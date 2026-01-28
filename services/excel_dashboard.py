"""
Executive Excel Dashboard Generator - 4-Sheet Professional Format
Industry-grade formatting with openpyxl for Municipal Pipe Tool Analysis Results

Sheet 1: Executive Summary - Document overview, statistics, visual breakdown
Sheet 2: Detailed Results - All questions with answers in unified table
Sheet 3: By Section - Questions grouped by section with clear headers
Sheet 4: Footnotes - All footnotes and citation references
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelDashboardGenerator:
    """Generate executive-ready Excel dashboards with 4 professional sheets"""

    # Professional color scheme - Municipal Pipe Tool Purple branding
    PURPLE_DARK = "667EEA"
    PURPLE = "764BA2"
    PURPLE_LIGHT = "E9E4F0"
    GREEN = "22C55E"
    GREEN_LIGHT = "DCFCE7"
    RED = "EF4444"
    RED_LIGHT = "FEE2E2"
    ORANGE = "F59E0B"
    ORANGE_LIGHT = "FEF3C7"
    GRAY = "6B7280"
    GRAY_LIGHT = "F3F4F6"
    WHITE = "FFFFFF"

    # Fills
    HEADER_FILL = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    SUBHEADER_FILL = PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid")
    SECTION_FILL = PatternFill(start_color="E9E4F0", end_color="E9E4F0", fill_type="solid")
    ANSWERED_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    UNANSWERED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    ALT_ROW_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    STAT_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    TITLE_FILL = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")

    # Fonts
    TITLE_FONT = Font(name='Calibri', size=24, bold=True, color="FFFFFF")
    HEADER_FONT = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    SUBHEADER_FONT = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    SECTION_FONT = Font(name='Calibri', size=12, bold=True, color="667EEA")
    DATA_FONT = Font(name='Calibri', size=11, color="374151")
    DATA_FONT_BOLD = Font(name='Calibri', size=11, bold=True, color="1F2937")
    STAT_LABEL_FONT = Font(name='Calibri', size=11, color="6B7280")
    STAT_VALUE_FONT = Font(name='Calibri', size=16, bold=True, color="667EEA")
    LINK_FONT = Font(name='Calibri', size=11, color="764BA2", underline='single')

    # Borders
    BORDER_THIN = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    BORDER_MEDIUM = Border(
        left=Side(style='medium', color='9CA3AF'),
        right=Side(style='medium', color='9CA3AF'),
        top=Side(style='medium', color='9CA3AF'),
        bottom=Side(style='medium', color='9CA3AF')
    )

    def __init__(self, analysis_result, is_partial=False):
        """
        Initialize dashboard generator

        Args:
            analysis_result: Dict with structure:
                {
                    'sections': [
                        {
                            'section_name': str,
                            'questions': [
                                {'question': str, 'answer': str, 'page_citations': [int], 'footnote': str}
                            ]
                        }
                    ],
                    'document_name': str (optional)
                }
            is_partial: Boolean flag indicating if this is a partial/stopped analysis
        """
        self.result = analysis_result
        self.is_partial = is_partial
        self.wb = Workbook()
        self.footnotes = self._collect_footnotes()

    def generate(self):
        """Generate complete 4-sheet dashboard workbook"""
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        # Create all 4 sheets
        self._create_executive_summary()      # Sheet 1: Executive Summary
        self._create_detailed_results()       # Sheet 2: Detailed Results
        self._create_by_section()             # Sheet 3: By Section
        self._create_footnotes_sheet()        # Sheet 4: Footnotes

        # Save to bytes
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output

    def _collect_footnotes(self):
        """Collect all footnotes from the analysis"""
        footnotes = []
        fn_num = 1
        for section in self.result.get('sections', []):
            for q in section.get('questions', []):
                fn = q.get('footnote')
                if fn and fn.strip():
                    footnotes.append({
                        'number': fn_num,
                        'section': section.get('section_name', 'Unknown'),
                        'question': q.get('question', ''),
                        'footnote': fn.strip(),
                        'pages': q.get('page_citations', [])
                    })
                    fn_num += 1
        return footnotes

    def _create_executive_summary(self):
        """Sheet 1: Executive Summary - Professional overview with statistics"""
        ws = self.wb.create_sheet('Executive Summary', 0)

        # Calculate statistics
        stats = self._calculate_statistics()
        section_stats = self._calculate_section_stats()

        # === TITLE BANNER ===
        ws.merge_cells('A1:F2')
        title_cell = ws['A1']
        title_cell.value = 'MUNICIPAL PIPE TOOL - ANALYSIS REPORT'
        title_cell.font = self.TITLE_FONT
        title_cell.fill = self.TITLE_FILL
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 25

        # Subtitle row
        ws.merge_cells('A3:F3')
        ws['A3'] = f'Generated: {self._get_timestamp()}'
        ws['A3'].font = Font(name='Calibri', size=11, italic=True, color="6B7280")
        ws['A3'].alignment = Alignment(horizontal='center')

        row = 4

        # === PARTIAL RESULTS BANNER ===
        if self.is_partial:
            row += 1
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = '⚠ PARTIAL RESULTS - Analysis was stopped before completion'
            ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color="92400E")
            ws[f'A{row}'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row].height = 30

        # === KEY METRICS SECTION ===
        row += 2
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = 'KEY METRICS'
        ws[f'A{row}'].font = self.SECTION_FONT
        ws[f'A{row}'].fill = self.SECTION_FILL
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 28

        row += 1
        # Create 4 metric boxes in a row
        metrics = [
            ('Total Questions', str(stats['total']), self.PURPLE_DARK),
            ('Answered', str(stats['answered']), self.GREEN),
            ('Not Found', str(stats['unanswered']), self.RED),
            ('Answer Rate', f"{stats['answer_rate']:.1f}%", self.PURPLE)
        ]

        col = 1
        for label, value, color in metrics:
            # Metric box - label
            cell = ws.cell(row, col, label)
            cell.font = self.STAT_LABEL_FONT
            cell.alignment = Alignment(horizontal='center')

            # Metric box - value
            value_cell = ws.cell(row + 1, col, value)
            value_cell.font = Font(name='Calibri', size=20, bold=True, color=color)
            value_cell.alignment = Alignment(horizontal='center')

            col += 1

        ws.row_dimensions[row].height = 20
        ws.row_dimensions[row + 1].height = 35
        row += 3

        # === SECTION BREAKDOWN TABLE ===
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = 'ANALYSIS BY SECTION'
        ws[f'A{row}'].font = self.SECTION_FONT
        ws[f'A{row}'].fill = self.SECTION_FILL
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 28

        row += 1
        # Table headers
        headers = ['Section Name', 'Questions', 'Answered', 'Not Found', 'Rate', 'Status']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER_THIN
        ws.row_dimensions[row].height = 25

        # Table data rows
        for idx, sec in enumerate(section_stats):
            row += 1
            is_alt = idx % 2 == 1

            # Section name
            ws.cell(row, 1, sec['name']).font = self.DATA_FONT_BOLD
            ws.cell(row, 2, sec['total']).font = self.DATA_FONT
            ws.cell(row, 3, sec['answered']).font = self.DATA_FONT
            ws.cell(row, 4, sec['unanswered']).font = self.DATA_FONT
            ws.cell(row, 5, f"{sec['rate']:.0f}%").font = self.DATA_FONT_BOLD

            # Status indicator
            if sec['rate'] == 100:
                status = 'Complete'
                status_fill = self.ANSWERED_FILL
            elif sec['rate'] >= 75:
                status = 'Good'
                status_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            elif sec['rate'] >= 50:
                status = 'Partial'
                status_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            else:
                status = 'Low'
                status_fill = self.UNANSWERED_FILL

            status_cell = ws.cell(row, 6, status)
            status_cell.font = Font(name='Calibri', size=10, bold=True)
            status_cell.fill = status_fill

            # Apply styling to all cells in row
            for col in range(1, 7):
                cell = ws.cell(row, col)
                cell.border = self.BORDER_THIN
                cell.alignment = Alignment(horizontal='left' if col == 1 else 'center', vertical='center')
                if is_alt and col != 6:
                    cell.fill = self.ALT_ROW_FILL

            ws.row_dimensions[row].height = 22

        # === TOTALS ROW ===
        row += 1
        ws.cell(row, 1, 'TOTAL').font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        ws.cell(row, 1).fill = self.SUBHEADER_FILL
        ws.cell(row, 2, stats['total']).font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        ws.cell(row, 2).fill = self.SUBHEADER_FILL
        ws.cell(row, 3, stats['answered']).font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        ws.cell(row, 3).fill = self.SUBHEADER_FILL
        ws.cell(row, 4, stats['unanswered']).font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        ws.cell(row, 4).fill = self.SUBHEADER_FILL
        ws.cell(row, 5, f"{stats['answer_rate']:.0f}%").font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        ws.cell(row, 5).fill = self.SUBHEADER_FILL
        ws.cell(row, 6, '').fill = self.SUBHEADER_FILL

        for col in range(1, 7):
            ws.cell(row, col).border = self.BORDER_THIN
            ws.cell(row, col).alignment = Alignment(horizontal='left' if col == 1 else 'center', vertical='center')

        # === FOOTNOTES SUMMARY ===
        if self.footnotes:
            row += 3
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = f'FOOTNOTES: {len(self.footnotes)} citations collected (see Footnotes sheet for details)'
            ws[f'A{row}'].font = Font(name='Calibri', size=11, italic=True, color="6B7280")

        # Column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12

    def _create_detailed_results(self):
        """Sheet 2: Detailed Results - All Q&A in unified professional table"""
        ws = self.wb.create_sheet('Detailed Results')

        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = 'COMPLETE ANALYSIS RESULTS'
        ws['A1'].font = Font(name='Calibri', size=16, bold=True, color="667EEA")
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30

        # Headers
        row = 2
        headers = ['#', 'Section', 'Question', 'Answer', 'PDF Pages', 'Footnote', 'Status']
        widths = [5, 25, 45, 60, 12, 8, 11]

        for col, (header, width) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row, col, header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.BORDER_THIN
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[row].height = 25

        # Data rows
        row = 3
        question_num = 1
        current_section = None

        for section in self.result.get('sections', []):
            section_name = section.get('section_name', '')

            for q in section.get('questions', []):
                answer_text = q.get('answer') or ''
                has_answer = bool(answer_text and answer_text.strip())
                pages = q.get('page_citations', [])
                pages_str = ', '.join(map(str, pages)) if pages else '-'
                has_footnote = bool(q.get('footnote'))

                # Determine row styling
                is_section_start = section_name != current_section
                is_alt_row = question_num % 2 == 0

                # Question number
                ws.cell(row, 1, question_num).font = self.DATA_FONT
                ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='top')

                # Section (show only on first question of each section)
                section_cell = ws.cell(row, 2, section_name if is_section_start else '')
                if is_section_start:
                    section_cell.font = self.DATA_FONT_BOLD
                    current_section = section_name
                else:
                    section_cell.font = self.DATA_FONT

                # Question
                ws.cell(row, 3, q.get('question', '')).font = self.DATA_FONT

                # Answer
                answer_cell = ws.cell(row, 4, answer_text if has_answer else 'Not found in document')
                if has_answer:
                    answer_cell.font = self.DATA_FONT
                else:
                    answer_cell.font = Font(name='Calibri', size=11, italic=True, color="9CA3AF")

                # PDF Pages
                pages_cell = ws.cell(row, 5, pages_str)
                pages_cell.font = self.DATA_FONT_BOLD if has_answer else self.DATA_FONT
                pages_cell.alignment = Alignment(horizontal='center', vertical='top')

                # Footnote indicator
                fn_cell = ws.cell(row, 6, '✓' if has_footnote else '-')
                fn_cell.font = Font(name='Calibri', size=11, bold=True, color=self.GREEN if has_footnote else self.GRAY)
                fn_cell.alignment = Alignment(horizontal='center', vertical='top')

                # Status
                status_cell = ws.cell(row, 7, 'Found' if has_answer else 'Not Found')
                status_cell.font = Font(name='Calibri', size=10, bold=True)
                status_cell.fill = self.ANSWERED_FILL if has_answer else self.UNANSWERED_FILL
                status_cell.alignment = Alignment(horizontal='center', vertical='top')

                # Apply borders and alternating row styling
                for col in range(1, 8):
                    cell = ws.cell(row, col)
                    cell.border = self.BORDER_THIN
                    if col not in [6, 7]:  # Don't override footnote/status fills
                        if is_alt_row:
                            cell.fill = self.ALT_ROW_FILL
                    cell.alignment = Alignment(
                        horizontal='left' if col in [2, 3, 4] else 'center',
                        vertical='top',
                        wrap_text=True
                    )

                ws.row_dimensions[row].height = 55
                row += 1
                question_num += 1

    def _create_by_section(self):
        """Sheet 3: By Section - Questions grouped with clear visual separation"""
        ws = self.wb.create_sheet('By Section')

        row = 1
        for section_idx, section in enumerate(self.result.get('sections', [])):
            # Section header with gradient effect
            ws.merge_cells(f'A{row}:F{row}')
            header_cell = ws[f'A{row}']
            header_cell.value = section.get('section_name', 'Unknown Section').upper()
            header_cell.font = Font(name='Calibri', size=13, bold=True, color="FFFFFF")
            header_cell.fill = self.HEADER_FILL
            header_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[row].height = 32
            row += 1

            # Column headers for this section
            headers = ['#', 'Question', 'Answer', 'PDF Pages', 'Notes', 'Status']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row, col, header)
                cell.font = self.SUBHEADER_FONT
                cell.fill = self.SUBHEADER_FILL
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = self.BORDER_THIN
            ws.row_dimensions[row].height = 22
            row += 1

            # Questions in this section
            for idx, q in enumerate(section.get('questions', []), start=1):
                answer_text = q.get('answer') or ''
                has_answer = bool(answer_text and answer_text.strip())
                pages = q.get('page_citations', [])
                pages_str = ', '.join(map(str, pages)) if pages else '-'
                is_alt = idx % 2 == 0

                # Row number
                ws.cell(row, 1, idx).font = self.DATA_FONT
                ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='top')

                # Question
                ws.cell(row, 2, q.get('question', '')).font = self.DATA_FONT

                # Answer
                answer_cell = ws.cell(row, 3, answer_text if has_answer else 'Not found in document')
                if has_answer:
                    answer_cell.font = self.DATA_FONT
                else:
                    answer_cell.font = Font(name='Calibri', size=11, italic=True, color="9CA3AF")

                # PDF Pages
                ws.cell(row, 4, pages_str).font = self.DATA_FONT_BOLD if has_answer else self.DATA_FONT
                ws.cell(row, 4).alignment = Alignment(horizontal='center', vertical='top')

                # Notes (footnote indicator)
                has_footnote = bool(q.get('footnote'))
                notes_cell = ws.cell(row, 5, 'See footnotes' if has_footnote else '-')
                notes_cell.font = Font(name='Calibri', size=10, italic=True, color=self.PURPLE if has_footnote else self.GRAY)
                notes_cell.alignment = Alignment(horizontal='center', vertical='top')

                # Status
                status_cell = ws.cell(row, 6, '✓' if has_answer else '✗')
                status_cell.font = Font(name='Calibri', size=14, bold=True, color=self.GREEN if has_answer else self.RED)
                status_cell.fill = self.ANSWERED_FILL if has_answer else self.UNANSWERED_FILL
                status_cell.alignment = Alignment(horizontal='center', vertical='top')

                # Apply borders and alternating styling
                for col in range(1, 7):
                    cell = ws.cell(row, col)
                    cell.border = self.BORDER_THIN
                    if col != 6 and is_alt:
                        cell.fill = self.ALT_ROW_FILL
                    cell.alignment = Alignment(
                        horizontal='left' if col in [2, 3] else 'center',
                        vertical='top',
                        wrap_text=True
                    )

                ws.row_dimensions[row].height = 50
                row += 1

            # Section spacing
            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 10

    def _create_footnotes_sheet(self):
        """Sheet 4: Footnotes - All citations and contextual notes"""
        ws = self.wb.create_sheet('Footnotes')

        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = 'FOOTNOTES & CITATIONS'
        ws['A1'].font = Font(name='Calibri', size=16, bold=True, color="667EEA")
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30

        if not self.footnotes:
            ws['A3'] = 'No footnotes were generated during this analysis.'
            ws['A3'].font = Font(name='Calibri', size=11, italic=True, color="6B7280")
            return

        # Summary
        ws['A2'] = f'{len(self.footnotes)} footnotes collected from analysis'
        ws['A2'].font = Font(name='Calibri', size=11, italic=True, color="6B7280")

        # Headers
        row = 4
        headers = ['#', 'Section', 'Related Question', 'Footnote/Citation', 'PDF Pages']
        widths = [5, 25, 40, 55, 12]

        for col, (header, width) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row, col, header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.BORDER_THIN
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.row_dimensions[row].height = 25

        # Footnote rows
        for fn in self.footnotes:
            row += 1
            is_alt = fn['number'] % 2 == 0

            ws.cell(row, 1, fn['number']).font = self.DATA_FONT_BOLD
            ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='top')

            ws.cell(row, 2, fn['section']).font = self.DATA_FONT
            ws.cell(row, 3, fn['question'][:100] + '...' if len(fn['question']) > 100 else fn['question']).font = self.DATA_FONT
            ws.cell(row, 4, fn['footnote']).font = self.DATA_FONT

            pages_str = ', '.join(map(str, fn['pages'])) if fn['pages'] else '-'
            ws.cell(row, 5, pages_str).font = self.DATA_FONT_BOLD
            ws.cell(row, 5).alignment = Alignment(horizontal='center', vertical='top')

            # Apply borders and alternating styling
            for col in range(1, 6):
                cell = ws.cell(row, col)
                cell.border = self.BORDER_THIN
                if is_alt:
                    cell.fill = self.ALT_ROW_FILL
                cell.alignment = Alignment(
                    horizontal='left' if col in [2, 3, 4] else 'center',
                    vertical='top',
                    wrap_text=True
                )

            ws.row_dimensions[row].height = 45

    def _calculate_statistics(self):
        """Calculate overall statistics"""
        all_questions = []
        for section in self.result.get('sections', []):
            all_questions.extend(section.get('questions', []))

        total = len(all_questions)
        answered = sum(1 for q in all_questions if q.get('answer'))
        unanswered = total - answered
        answer_rate = (answered / total * 100) if total > 0 else 0

        return {
            'total': total,
            'answered': answered,
            'unanswered': unanswered,
            'answer_rate': answer_rate
        }

    def _calculate_section_stats(self):
        """Calculate per-section statistics"""
        section_stats = []
        for section in self.result.get('sections', []):
            questions = section.get('questions', [])
            total = len(questions)
            answered = sum(1 for q in questions if q.get('answer'))
            unanswered = total - answered
            rate = (answered / total * 100) if total > 0 else 0

            section_stats.append({
                'name': section.get('section_name', 'Unknown'),
                'total': total,
                'answered': answered,
                'unanswered': unanswered,
                'rate': rate
            })
        return section_stats

    def _get_timestamp(self):
        """Get formatted timestamp"""
        return datetime.now().strftime('%B %d, %Y at %I:%M %p')
