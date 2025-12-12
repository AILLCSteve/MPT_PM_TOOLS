"""
Visual Project Summary - Plotly Dash Application
Complete feature-rich dashboard with all visualizations and tables
Integrated into main Flask app as a sub-application
"""

import dash
from dash import dcc, html, dash_table, Input, Output, State
import dash_bootstrap_components as dbc
from dash.exceptions import PreventUpdate
import plotly.graph_objects as go
import plotly.express as px
import pandas as pd
import base64
import io
import os
import threading
from datetime import datetime
from pathlib import Path

from services.cipp_dashboard.data_processor import CIPPDataProcessor
from services.cipp_dashboard.excel_generator_v2 import ExcelDashboardGeneratorV2

# Color scheme - vibrant, distinct colors for lifecycle stages
COLORS = {
    'Not Started': '#E0E0E0',  # Grey (not yet in pipeline)
    'Awaiting Prep': '#FF6B35',  # Vibrant orange
    'Ready to Line': '#F7B801',  # Vibrant yellow/gold
    'Wet Out': '#3498DB',  # Bright blue
    'Lined': '#004E89',  # Deep blue
    'Post TV Complete': '#6A0572'  # Deep purple
}

STAGE_ORDER = ['Not Started', 'Awaiting Prep', 'Ready to Line', 'Wet Out', 'Lined', 'Post TV Complete']
LIFECYCLE_STAGES = ['Awaiting Prep', 'Ready to Line', 'Wet Out', 'Lined', 'Post TV Complete']  # Exclude Not Started


def create_dash_app(flask_app):
    """Create and configure Dash app to work with Flask"""

    # Initialize Dash app with Bootstrap theme and integrate with Flask
    dash_app = dash.Dash(
        __name__,
        server=flask_app,
        url_base_pathname='/cipp-dashboard/',
        external_stylesheets=[dbc.themes.BOOTSTRAP, dbc.icons.FONT_AWESOME],
        suppress_callback_exceptions=True,
        meta_tags=[
            {"name": "viewport", "content": "width=device-width, initial-scale=1.0, maximum-scale=5.0, user-scalable=yes"}
        ]
    )

    dash_app.title = "Visual Project Summary"

    # Create folders (use absolute paths relative to project root)
    base_dir = Path(__file__).parent
    uploads_dir = base_dir / 'uploads'
    outputs_dir = base_dir / 'outputs'
    uploads_dir.mkdir(exist_ok=True)
    outputs_dir.mkdir(exist_ok=True)

    # Cache for processors (per-worker, keyed by filepath not session)
    # This is safe because each worker process has its own cache
    processor_cache = {}

    def get_processor(filepath):
        """Get or create processor from filepath (with per-worker caching)"""
        if not filepath or not os.path.exists(filepath):
            return None

        # Cache by filepath so multiple callbacks reuse same processor
        if filepath not in processor_cache:
            processor = CIPPDataProcessor(filepath)
            processor.load_data()
            processor_cache[filepath] = processor

        return processor_cache[filepath]

    # Layout
    dash_app.layout = dbc.Container(fluid=True, className="px-2 px-md-4", children=[
        # Header
        dbc.Navbar(
            dbc.Container([
                html.A(
                    dbc.Row([
                        dbc.Col(html.I(className="fas fa-chart-line fa-2x me-3")),
                        dbc.Col(dbc.NavbarBrand("Visual Project Summary", className="fs-3 fw-bold")),
                    ], align="center"),
                    href="/",
                    style={"textDecoration": "none"}
                )
            ]),
            color="primary",
            dark=True,
            className="mb-4"
        ),

        # Upload Section
        dbc.Row([
            dbc.Col([
                dbc.Card([
                    dbc.CardHeader(html.H5([html.I(className="fas fa-cloud-upload-alt me-2"), "Upload Project Schedule"])),
                    dbc.CardBody([
                        dcc.Upload(
                            id='upload-data',
                            children=html.Div([
                                html.I(className="fas fa-file-excel fa-3x mb-3 text-success"),
                                html.H5('Drag and Drop or Click to Select Excel File'),
                                html.P('.xlsx files only', className='text-muted')
                            ], className='text-center p-5 border border-2 border-dashed rounded'),
                            multiple=False,
                            accept='.xlsx'
                        ),
                        html.Div(id='upload-status', className='mt-3')
                    ])
                ], className='shadow')
            ], width=12)
        ], className='mb-4', id='upload-section'),

        # Dashboard Section (hidden initially)
        html.Div([
            # KPIs
            dbc.Row([
                dbc.Col([
                    dbc.Card([
                        dbc.CardBody([
                            html.I(className="fas fa-list fa-2x text-primary mb-2"),
                            html.H3(id='kpi-segments', className='fw-bold'),
                            html.P('Total Segments', className='text-muted mb-0')
                        ], className='text-center')
                    ], className='shadow-sm h-100')
                ], xs=12, sm=6, lg=2),
                dbc.Col([
                    dbc.Card([
                        dbc.CardBody([
                            html.I(className="fas fa-clipboard-check fa-2x text-info mb-2"),
                            html.H3(id='kpi-ready-to-line', className='fw-bold'),
                            html.P('Ready to Line', className='text-muted mb-0')
                        ], className='text-center')
                    ], className='shadow-sm h-100')
                ], xs=12, sm=6, lg=2),
                dbc.Col([
                    dbc.Card([
                        dbc.CardBody([
                            html.I(className="fas fa-ruler fa-2x text-success mb-2"),
                            html.H3(id='kpi-footage', className='fw-bold'),
                            html.P('Total Footage', className='text-muted mb-0')
                        ], className='text-center')
                    ], className='shadow-sm h-100')
                ], xs=12, sm=6, lg=2),
                dbc.Col([
                    dbc.Card([
                        dbc.CardBody([
                            html.I(className="fas fa-calculator fa-2x text-warning mb-2"),
                            html.H3(id='kpi-avg-length', className='fw-bold'),
                            html.P('Avg Length (ft)', className='text-muted mb-0')
                        ], className='text-center')
                    ], className='shadow-sm h-100')
                ], xs=12, sm=6, lg=2),
                dbc.Col([
                    html.Div([
                        dbc.Card([
                            dbc.CardBody([
                                html.I(className="fas fa-hard-hat fa-2x text-secondary mb-2"),
                                html.H3(id='kpi-prep-complete', className='fw-bold'),
                                html.P('% Prep Complete', className='text-muted mb-0')
                            ], className='text-center')
                        ], className='shadow-sm h-100')
                    ], id='kpi-prep-complete-card', style={'cursor': 'pointer', 'height': '100%'})
                ], xs=12, sm=6, lg=2),
                dbc.Col([
                    html.Div([
                        dbc.Card([
                            dbc.CardBody([
                                html.I(className="fas fa-tasks fa-2x text-danger mb-2"),
                                html.H3(id='kpi-completion', className='fw-bold'),
                                html.P('% CIPP Install Completion', className='text-muted mb-0')
                            ], className='text-center')
                        ], className='shadow-sm h-100')
                    ], id='kpi-completion-card', style={'cursor': 'pointer', 'height': '100%'})
                ], xs=12, sm=6, lg=2)
            ], className='mb-4'),

            # Hidden stores for toggle state
            dcc.Store(id='prep-toggle-state', data={'show_fraction': False}),
            dcc.Store(id='completion-toggle-state', data={'show_fraction': False}),

            # Download Buttons
            dbc.Row([
                dbc.Col([
                    dbc.Card([
                        dbc.CardBody([
                            html.H6([html.I(className="fas fa-download me-2"), "Download Excel Files"], className='mb-3'),
                            dbc.ButtonGroup([
                                dbc.Button([html.I(className="fas fa-file-excel me-2"), "Approach 1: Native Charts"],
                                          id='download-btn-1', color='primary', outline=True, className='me-2 mb-2 mb-md-0'),
                                dbc.Button([html.I(className="fas fa-file-excel me-2"), "Approach 2: Enhanced"],
                                          id='download-btn-2', color='success', outline=True, className='me-2 mb-2 mb-md-0'),
                                dbc.Button([html.I(className="fas fa-file-excel me-2"), "Approach 3: Plotly Images"],
                                          id='download-btn-3', color='warning', outline=True, className='mb-2 mb-md-0')
                            ], className='w-100 flex-column flex-md-row'),
                            dcc.Download(id="download-file")
                        ])
                    ], className='shadow-sm')
                ], width=12)
            ], className='mb-4'),

            # Visualizations
            # Overall Progress Bar (Completed vs Uncompleted)
            dbc.Row([
                dbc.Col([
                    dbc.Card([
                        dbc.CardHeader(html.H6("Overall Project Progress")),
                        dbc.CardBody([
                            dcc.Graph(id='overall-progress-chart', config={'displayModeBar': False, 'responsive': True})
                        ])
                    ], className='shadow-sm')
                ], width=12)
            ], className='mb-4'),

            # Breakout Tables - Interactive Filtered Segment Views
            dbc.Row([
                dbc.Col([
                    dbc.Card([
                        dbc.CardHeader(html.H5([
                            html.I(className="fas fa-filter me-2"),
                            "Breakout Tables - Filtered Segment Views"
                        ])),
                        dbc.CardBody([
                            html.P([
                                "Click on any chart above to filter segments, or use the tabs below for instant breakouts. ",
                                "Tables show segments matching specific criteria."
                            ], className='text-muted mb-3'),
                            dbc.Tabs([
                                dbc.Tab(label="Awaiting Prep", tab_id="breakout-awaiting"),
                                dbc.Tab(label="Ready to Line", tab_id="breakout-ready"),
                                dbc.Tab(label="CCTV Posted", tab_id="breakout-cctv"),
                                dbc.Tab(label="Pending", tab_id="breakout-pending"),
                                dbc.Tab(label="Easement", tab_id="breakout-easement"),
                                dbc.Tab(label="Traffic Control", tab_id="breakout-traffic"),
                                dbc.Tab(label="ROW Only", tab_id="breakout-row"),
                                dbc.Tab(label="All Segments", tab_id="breakout-all"),
                            ], id="breakout-tabs", active_tab="breakout-awaiting"),
                            html.Div(id='breakout-table-content', className='mt-3')
                        ])
                    ], className='shadow-sm')
                ], width=12)
            ], className='mb-4'),

            # Stage Progress
            dbc.Row([
                dbc.Col([
                    dbc.Card([
                        dbc.CardHeader(html.H6("Stage Progress")),
                        dbc.CardBody([
                            dcc.Graph(id='progress-bar-chart', config={'displayModeBar': False, 'responsive': True})
                        ])
                    ], className='shadow-sm')
                ], width=12)
            ], className='mb-4'),

            # Segment Characteristics (full width for radial chart)
            dbc.Row([
                dbc.Col([
                    dbc.Card([
                        dbc.CardHeader(html.H6("Segment Characteristics")),
                        dbc.CardBody([
                            dcc.Graph(id='radar-chart', config={'displayModeBar': False, 'responsive': True})
                        ])
                    ], className='shadow-sm')
                ], width=12)
            ], className='mb-4'),

            dbc.Row([
                dbc.Col([
                    dbc.Card([
                        dbc.CardHeader(html.H6("Progress by Pipe Size")),
                        dbc.CardBody([
                            dcc.Graph(id='pipe-progress-chart')
                        ])
                    ], className='shadow-sm h-100')
                ], md=6),
                dbc.Col([
                    dbc.Card([
                        dbc.CardHeader(html.H6("Pipe Size Distribution")),
                        dbc.CardBody([
                            dcc.Graph(id='pipe-size-chart')
                        ])
                    ], className='shadow-sm h-100')
                ], md=6)
            ], className='mb-4'),

            dbc.Row([
                dbc.Col([
                    dbc.Card([
                        dbc.CardHeader(html.H6("Segment Length Distribution")),
                        dbc.CardBody([
                            dcc.Graph(id='length-distribution-chart')
                        ])
                    ], className='shadow-sm h-100')
                ], md=6),
                dbc.Col([
                    dbc.Card([
                        dbc.CardHeader(html.H6("Segment Type Distribution")),
                        dbc.CardBody([
                            dcc.Graph(id='easement-traffic-chart')
                        ])
                    ], className='shadow-sm h-100')
                ], md=6)
            ], className='mb-4'),

            # Data Tables
            dbc.Row([
                dbc.Col([
                    dbc.Card([
                        dbc.CardHeader(html.H5([html.I(className="fas fa-table me-2"), "Summary Tables"])),
                        dbc.CardBody([
                            dbc.Tabs([
                                dbc.Tab(label="Stage Summary", tab_id="tab-1"),
                                dbc.Tab(label="Stage by Pipe Size", tab_id="tab-2"),
                                dbc.Tab(label="Pipe Size Mix", tab_id="tab-3"),
                                dbc.Tab(label="Length Bins", tab_id="tab-4"),
                                dbc.Tab(label="Easement/Traffic", tab_id="tab-5"),
                            ], id="table-tabs", active_tab="tab-1"),
                            html.Div(id='table-content', className='mt-3')
                        ])
                    ], className='shadow-sm')
                ], width=12)
            ], className='mb-4'),

            # Original Excel File Embedded
            dbc.Row([
                dbc.Col([
                    dbc.Card([
                        dbc.CardHeader(html.H5([html.I(className="fas fa-file-excel me-2"), "Original Excel Data"])),
                        dbc.CardBody([
                            html.Div(id='excel-table-container')
                        ])
                    ], className='shadow-sm')
                ], width=12)
            ], className='mb-5'),

        ], id='dashboard-section', style={'display': 'none'}),

        # Store session data
        dcc.Store(id='session-data'),

        # Store for chart click navigation to breakout tables
        dcc.Store(id='breakout-nav-store'),

        # Hidden signal that fires ONLY when data is fully uploaded and ready
        html.Div(id='data-ready-signal', style={'display': 'none'}),

    ])


    # ============================================================================
    # CALLBACKS - ALL FEATURES FROM ORIGINAL APP
    # ============================================================================

    # Callback for file upload
    @dash_app.callback(
        [Output('upload-status', 'children'),
         Output('dashboard-section', 'style'),
         Output('session-data', 'data'),
         Output('kpi-segments', 'children'),
         Output('kpi-ready-to-line', 'children'),
         Output('kpi-footage', 'children'),
         Output('kpi-avg-length', 'children'),
         Output('data-ready-signal', 'children')],  # Signal sent LAST
        [Input('upload-data', 'contents')],
        [State('upload-data', 'filename')]
    )
    def process_upload(contents, filename):
        if contents is None:
            raise PreventUpdate

        try:
            # Decode uploaded file
            content_type, content_string = contents.split(',')
            decoded = base64.b64decode(content_string)

            # Save file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filepath = uploads_dir / f"{timestamp}_{filename}"

            with open(filepath, 'wb') as f:
                f.write(decoded)

            # Process data
            processor = CIPPDataProcessor(str(filepath))
            processor.load_data()

            # Get Ready to Line count for KPI
            tables = processor.get_all_tables()
            stage_summary = tables['stage_footage_summary']

            ready_to_line_count = sum(
                r['Segment_Count'] for r in stage_summary
                if r['Stage'] == 'Ready to Line'
            )

            # Create session data payload with filepath (stateless - works across workers)
            session_id = timestamp
            session_payload = {
                'session_id': session_id,
                'filepath': str(filepath),
                'filename': filename,
                'total_footage': processor.total_footage,
                'segment_count': len(processor.segments)
            }

            status = dbc.Alert([
                html.I(className="fas fa-check-circle me-2"),
                f"Success! Processed {len(processor.segments)} segments ({processor.total_footage:,.0f} feet)"
            ], color='success')

            # Return session payload with filepath - callbacks will reload from disk
            return (
                status,
                {'display': 'block'},
                session_payload,
                f"{len(processor.segments)}",
                f"{ready_to_line_count}",
                f"{processor.total_footage:,.0f} ft",
                f"{processor.total_footage / len(processor.segments):.1f} ft",
                f"ready-{session_id}"  # Signal: data is ready for this session
            )

        except Exception as e:
            error = dbc.Alert([
                html.I(className="fas fa-exclamation-triangle me-2"),
                f"Error: {str(e)}"
            ], color='danger')
            return error, {'display': 'none'}, {}, '', '', '', '', ''


    # Callback for overall progress bar (project lifecycle filling up + lining completion)
    @dash_app.callback(
        Output('overall-progress-chart', 'figure'),
        [Input('session-data', 'data'),
         Input('data-ready-signal', 'children')],  # Waits for ready signal
        prevent_initial_call=True
    )
    def update_overall_progress(session_data, ready_signal):
        if not session_data or not ready_signal:
            # Return empty figure instead of PreventUpdate to ensure component renders
            return go.Figure()

        # Reload processor from filepath (stateless - works across workers)
        filepath = session_data.get('filepath')
        processor = get_processor(filepath)
        if not processor:
            # Data not yet available, return empty figure
            return go.Figure()

        # Use MUTUALLY EXCLUSIVE stage summary for overall progress bar (no overlapping)
        stage_summary = processor.get_overall_progress_summary()
        total_footage = processor.total_footage
        total_segments = len(processor.segments)

        # Build stage data map
        stage_data_map = {s['Stage']: s for s in stage_summary}

        fig = go.Figure()

        # === FIRST BAR: Project Lifecycle ===
        # Add lifecycle stages first (vibrant colors)
        for stage in LIFECYCLE_STAGES:
            stage_info = stage_data_map.get(stage, {'Total_Feet': 0, 'Pct_of_Total_Feet': 0})
            pct = stage_info['Pct_of_Total_Feet'] * 100
            footage = stage_info['Total_Feet']

            if pct > 0:
                # Only show text if percentage is large enough
                text_display = f"{stage}<br>{pct:.1f}%" if pct >= 5 else ""

                fig.add_trace(go.Bar(
                    y=['Project Lifecycle'],
                    x=[pct],
                    name=stage,
                    orientation='h',
                    marker_color=COLORS[stage],
                    text=text_display,
                    textposition='inside',
                    textfont=dict(size=11, color='white', family='Arial', weight='bold'),
                    hovertemplate=f'<b>{stage}</b><br>{pct:.1f}% ({footage:,.0f} ft)<extra></extra>',
                    legendgroup='lifecycle'
                ))

        # Add "Not Started" as grey at the end
        not_started_info = stage_data_map.get('Not Started', {'Total_Feet': 0, 'Pct_of_Total_Feet': 0})
        not_started_pct = not_started_info['Pct_of_Total_Feet'] * 100
        not_started_footage = not_started_info['Total_Feet']

        if not_started_pct > 0:
            text_display = f"Not Yet Started<br>{not_started_pct:.1f}%" if not_started_pct >= 5 else ""

            fig.add_trace(go.Bar(
                y=['Project Lifecycle'],
                x=[not_started_pct],
                name='Not Yet Started',
                orientation='h',
                marker_color=COLORS['Not Started'],
                text=text_display,
                textposition='inside',
                textfont=dict(size=11, color='#666', family='Arial'),
                hovertemplate=f'<b>Not Yet Started</b><br>{not_started_pct:.1f}% ({not_started_footage:,.0f} ft)<extra></extra>',
                legendgroup='lifecycle'
            ))

        # === SECOND BAR: CIPP Lining Completion (by segment count) ===
        # Calculate segments with lining completed (Lined + Post TV Complete)
        lining_complete_stages = ['Lined', 'Post TV Complete']
        lining_complete_count = sum(
            stage_data_map.get(stage, {'Segment_Count': 0}).get('Segment_Count', 0)
            for stage in lining_complete_stages
        )
        lining_complete_pct = (lining_complete_count / total_segments * 100) if total_segments > 0 else 0
        lining_incomplete_pct = 100 - lining_complete_pct

        # Lining completed portion
        if lining_complete_pct > 0:
            text_display = f"Lining Complete<br>{lining_complete_pct:.1f}%" if lining_complete_pct >= 5 else ""
            fig.add_trace(go.Bar(
                y=['CIPP Lining Status'],
                x=[lining_complete_pct],
                name='Lining Complete',
                orientation='h',
                marker_color='#27AE60',  # Green
                text=text_display,
                textposition='inside',
                textfont=dict(size=11, color='white', family='Arial', weight='bold'),
                hovertemplate=f'<b>Lining Complete</b><br>{lining_complete_pct:.1f}% ({lining_complete_count} segments)<extra></extra>',
                legendgroup='lining',
                showlegend=False
            ))

        # Lining not yet completed portion
        if lining_incomplete_pct > 0:
            text_display = f"Lining Not Complete<br>{lining_incomplete_pct:.1f}%" if lining_incomplete_pct >= 5 else ""
            fig.add_trace(go.Bar(
                y=['CIPP Lining Status'],
                x=[lining_incomplete_pct],
                name='Lining Not Complete',
                orientation='h',
                marker_color='#E0E0E0',  # Grey
                text=text_display,
                textposition='inside',
                textfont=dict(size=11, color='#666', family='Arial'),
                hovertemplate=f'<b>Lining Not Complete</b><br>{lining_incomplete_pct:.1f}% ({total_segments - lining_complete_count} segments)<extra></extra>',
                legendgroup='lining',
                showlegend=False
            ))

        fig.update_layout(
            barmode='stack',
            showlegend=True,
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=-0.5,
                xanchor="center",
                x=0.5,
                font=dict(size=10),
                traceorder='normal'
            ),
            height=240,
            autosize=True,
            margin=dict(l=150, r=20, t=10, b=90),
            xaxis=dict(
                title="",
                showgrid=False,
                range=[0, 100],
                ticksuffix='%',
                fixedrange=True
            ),
            yaxis=dict(
                showticklabels=True,
                tickfont=dict(size=12, family='Arial', weight='bold'),
                categoryorder='array',
                categoryarray=['CIPP Lining Status', 'Project Lifecycle']
            ),
            plot_bgcolor='white',
            paper_bgcolor='white',
            font=dict(size=12)
        )

        return fig


    # Callback for stage progress bar (horizontal bars showing completed ft / total ft per stage)
    @dash_app.callback(
        Output('progress-bar-chart', 'figure'),
        [Input('session-data', 'data'),
         Input('data-ready-signal', 'children')],
        prevent_initial_call=True
    )
    def update_progress_bar(session_data, ready_signal):
        if not session_data or not ready_signal:
            return go.Figure()

        filepath = session_data.get('filepath')
        processor = get_processor(filepath)
        if not processor:
            return go.Figure()

        tables = processor.get_all_tables()
        stage_summary = tables['stage_footage_summary']
        total_footage = processor.total_footage

        # Build stage data map
        stage_data_map = {s['Stage']: s for s in stage_summary}

        fig = go.Figure()

        # Only show lifecycle stages (exclude Not Started)
        for stage in reversed(LIFECYCLE_STAGES):  # Reverse so Prep is at top
            stage_info = stage_data_map.get(stage, {'Total_Feet': 0})
            completed_ft = stage_info['Total_Feet']

            # Show completed ft / total ft for ALL stages
            if completed_ft > 0:
                fig.add_trace(go.Bar(
                    y=[stage],
                    x=[completed_ft],
                    name=stage,
                    orientation='h',
                    marker_color=COLORS[stage],
                    text=f"{completed_ft:,.0f} ft / {total_footage:,.0f} ft",
                    textposition='inside',
                    textfont=dict(size=14, color='white', family='Arial', weight='bold'),
                    hovertemplate=f'<b>{stage}</b><br>Completed: {completed_ft:,.0f} ft<br>Total Project: {total_footage:,.0f} ft<br>% of Total: {(completed_ft/total_footage*100):.1f}%<extra></extra>',
                    showlegend=False
                ))
            else:
                # Show empty bar with text
                fig.add_trace(go.Bar(
                    y=[stage],
                    x=[total_footage * 0.05],  # Slightly larger baseline for text visibility
                    name=stage,
                    orientation='h',
                    marker_color='#F0F0F0',
                    text=f"0 ft / {total_footage:,.0f} ft",
                    textposition='inside',
                    textfont=dict(size=14, color='#666', family='Arial', weight='bold'),
                    hovertemplate=f'<b>{stage}</b><br>Completed: 0 ft<br>Total Project: {total_footage:,.0f} ft<br>% of Total: 0%<extra></extra>',
                    showlegend=False
                ))

        fig.update_layout(
            barmode='overlay',
            showlegend=False,
            height=250,
            autosize=True,
            margin=dict(l=160, r=30, t=20, b=50),
            xaxis=dict(
                title=dict(
                    text="Footage Completed",
                    font=dict(size=14)
                ),
                showgrid=True,
                gridcolor='rgba(200, 200, 200, 0.3)',
                range=[0, total_footage],
                tickformat=',',
                tickfont=dict(size=13),
                fixedrange=True
            ),
            yaxis=dict(
                showticklabels=True,
                tickfont=dict(size=13, family='Arial', weight='bold'),
                fixedrange=True
            ),
            plot_bgcolor='white',
            paper_bgcolor='white',
            font=dict(size=13)
        )

        return fig


    # Callback for radial bar chart (segment characteristics)
    @dash_app.callback(
        Output('radar-chart', 'figure'),
        [Input('session-data', 'data'),
         Input('data-ready-signal', 'children')],
        prevent_initial_call=True
    )
    def update_radar_chart(session_data, ready_signal):
        if not session_data or not ready_signal:
            return go.Figure()

        filepath = session_data.get('filepath')
        processor = get_processor(filepath)
        if not processor:
            return go.Figure()

        segments = processor.segments

        total = len(segments)

        # Calculate percentages and footage
        easement_segs = [s for s in segments if s['easement']]
        traffic_segs = [s for s in segments if s['traffic_control']]
        regular_segs = [s for s in segments if not s['easement'] and not s['traffic_control']]
        large_pipe_segs = [s for s in segments if s['pipe_size'] and s['pipe_size'] > 12]
        small_pipe_segs = [s for s in segments if s['pipe_size'] and s['pipe_size'] <= 12]

        easement_pct = len(easement_segs) / total * 100
        traffic_pct = len(traffic_segs) / total * 100
        regular_pct = len(regular_segs) / total * 100
        large_pipe_pct = len(large_pipe_segs) / total * 100
        small_pipe_pct = len(small_pipe_segs) / total * 100

        easement_ft = sum(s['map_length'] for s in easement_segs)
        traffic_ft = sum(s['map_length'] for s in traffic_segs)
        regular_ft = sum(s['map_length'] for s in regular_segs)
        large_pipe_ft = sum(s['map_length'] for s in large_pipe_segs)
        small_pipe_ft = sum(s['map_length'] for s in small_pipe_segs)

        categories = [
            f'Easement - {easement_ft:,.0f} ft',
            f'Traffic Control - {traffic_ft:,.0f} ft',
            f'Regular - {regular_ft:,.0f} ft',
            f'Large Pipe (>12") - {large_pipe_ft:,.0f} ft',
            f'Small Pipe (≤12") - {small_pipe_ft:,.0f} ft'
        ]
        values = [easement_pct, traffic_pct, regular_pct, large_pipe_pct, small_pipe_pct]
        colors = ['#FFC000', '#E74C3C', '#95A5A6', '#3498DB', '#2ECC71']

        fig = go.Figure()

        # Create radial bar chart
        fig.add_trace(go.Barpolar(
            r=values,
            theta=categories,
            marker=dict(
                color=colors,
                line=dict(color='white', width=2)
            ),
            hovertemplate='<b>%{theta}</b><br>%{r:.1f}%<extra></extra>',
            name='Segment Characteristics'
        ))

        fig.update_layout(
            polar=dict(
                radialaxis=dict(
                    visible=True,
                    range=[0, 100],
                    showticklabels=True,
                    ticksuffix='%',
                    tickfont=dict(size=14),
                    tickangle=0,
                    gridcolor='rgba(128, 128, 128, 0.2)'
                ),
                angularaxis=dict(
                    showticklabels=True,
                    tickfont=dict(size=13)
                ),
                bgcolor='rgba(240, 240, 240, 0.1)'
            ),
            showlegend=False,
            height=400,
            autosize=True,
            margin=dict(l=80, r=80, t=40, b=40),
            paper_bgcolor='white',
            font=dict(size=12, color='#333')
        )

        return fig


    # Callback for pipe progress chart
    @dash_app.callback(
        Output('pipe-progress-chart', 'figure'),
        [Input('session-data', 'data'),
         Input('data-ready-signal', 'children')],
        prevent_initial_call=True
    )
    def update_pipe_progress(session_data, ready_signal):
        if not session_data or not ready_signal:
            return go.Figure()

        filepath = session_data.get('filepath')
        processor = get_processor(filepath)
        if not processor:
            return go.Figure()

        tables = processor.get_all_tables()
        stage_by_pipe = tables['stage_by_pipe_size']

        fig = go.Figure()

        # Reverse the lifecycle stages so they stack from bottom-up (Prep at bottom, Post TV at top)
        # Exclude Not Started from this chart
        for stage in reversed(LIFECYCLE_STAGES):
            fig.add_trace(go.Bar(
                name=stage,
                x=[str(r['Pipe Size']) for r in stage_by_pipe],
                y=[r.get(stage, 0) for r in stage_by_pipe],  # Use .get() to handle missing stages
                marker_color=COLORS.get(stage, '#95A5A6'),
                hovertemplate='<b>%{x}" pipe</b><br>' + stage + '<br>%{y:,.0f} ft<extra></extra>'
            ))

        fig.update_layout(
            barmode='stack',
            xaxis_title="Pipe Size (inches)",
            yaxis_title="Footage",
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=-0.35,
                xanchor="center",
                x=0.5,
                font=dict(size=10),
                traceorder='reversed'  # Reverse legend order to match visual
            ),
            height=420,
            margin=dict(l=50, r=20, t=40, b=110),
            plot_bgcolor='white',
            paper_bgcolor='white',
            font=dict(size=11)
        )

        return fig


    # Callback for pipe size distribution (donut chart)
    @dash_app.callback(
        Output('pipe-size-chart', 'figure'),
        [Input('session-data', 'data'),
         Input('data-ready-signal', 'children')],
        prevent_initial_call=True
    )
    def update_pipe_size_chart(session_data, ready_signal):
        if not session_data or not ready_signal:
            return go.Figure()

        filepath = session_data.get('filepath')
        processor = get_processor(filepath)
        if not processor:
            return go.Figure()

        tables = processor.get_all_tables()
        pipe_mix = tables['pipe_size_mix']

        labels = [f"{r['Pipe Size']}\"" for r in pipe_mix]
        values = [r['Total_Feet'] for r in pipe_mix]

        fig = go.Figure(data=[go.Pie(
            labels=labels,
            values=values,
            hole=0.4,
            marker=dict(colors=px.colors.qualitative.Set3),
            textinfo='label+percent',
            textposition='outside',
            hovertemplate='<b>%{label}</b><br>%{value:,.0f} ft<br>%{percent}<extra></extra>'
        )])

        fig.update_layout(
            annotations=[dict(text='Pipe<br>Sizes', x=0.5, y=0.5, font_size=14, showarrow=False)],
            height=400,
            margin=dict(l=20, r=20, t=40, b=40),
            showlegend=True,
            legend=dict(orientation="v", yanchor="middle", y=0.5, xanchor="left", x=1.1),
            paper_bgcolor='white',
            font=dict(size=11)
        )

        return fig


    # Callback for length distribution (horizontal bar chart)
    @dash_app.callback(
        Output('length-distribution-chart', 'figure'),
        [Input('session-data', 'data'),
         Input('data-ready-signal', 'children')],
        prevent_initial_call=True
    )
    def update_length_distribution(session_data, ready_signal):
        if not session_data or not ready_signal:
            return go.Figure()

        filepath = session_data.get('filepath')
        processor = get_processor(filepath)
        if not processor:
            return go.Figure()

        tables = processor.get_all_tables()
        length_bins = tables['length_bins']

        # Create horizontal bar chart
        fig = go.Figure()

        fig.add_trace(go.Bar(
            y=[r['Length_Bin_Label'] for r in length_bins],
            x=[r['Total_Feet'] for r in length_bins],
            orientation='h',
            marker=dict(
                color=[r['Total_Feet'] for r in length_bins],
                colorscale='Viridis',
                showscale=False
            ),
            text=[f"{r['Segment_Count']} segments<br>{r['Total_Feet']:,.0f} ft" for r in length_bins],
            textposition='auto',
            hovertemplate='<b>%{y}</b><br>Footage: %{x:,.0f} ft<br>Segments: %{customdata}<extra></extra>',
            customdata=[r['Segment_Count'] for r in length_bins]
        ))

        fig.update_layout(
            xaxis_title="Total Footage",
            yaxis_title="Length Range (feet)",
            height=400,
            margin=dict(l=120, r=20, t=40, b=50),
            plot_bgcolor='white',
            paper_bgcolor='white',
            yaxis=dict(categoryorder='array', categoryarray=[r['Length_Bin_Label'] for r in reversed(length_bins)]),
            font=dict(size=11),
            showlegend=False
        )

        return fig


    # Callback for easement/traffic/regular segment types (pie chart)
    @dash_app.callback(
        Output('easement-traffic-chart', 'figure'),
        [Input('session-data', 'data'),
         Input('data-ready-signal', 'children')],
        prevent_initial_call=True
    )
    def update_easement_traffic(session_data, ready_signal):
        if not session_data or not ready_signal:
            return go.Figure()

        filepath = session_data.get('filepath')
        processor = get_processor(filepath)
        if not processor:
            return go.Figure()

        segments = processor.segments
        total_footage = processor.total_footage

        # Calculate footage for each category
        easement_footage = sum(s['map_length'] for s in segments if s['easement'])
        traffic_footage = sum(s['map_length'] for s in segments if s['traffic_control'])
        regular_footage = sum(s['map_length'] for s in segments if not s['easement'] and not s['traffic_control'])

        # Create labels and values
        labels = ['Easement', 'Traffic Control', 'Regular']
        values = [easement_footage, traffic_footage, regular_footage]
        colors = ['#FFC000', '#E74C3C', '#95A5A6']

        fig = go.Figure(data=[go.Pie(
            labels=labels,
            values=values,
            marker=dict(colors=colors),
            textinfo='label+percent',
            textposition='outside',
            textfont=dict(size=12),
            hovertemplate='<b>%{label}</b><br>%{value:,.0f} ft<br>%{percent}<extra></extra>',
            pull=[0.05, 0.05, 0],  # Slightly pull out easement and traffic slices
        )])

        fig.update_layout(
            height=400,
            margin=dict(l=20, r=20, t=40, b=40),
            showlegend=True,
            legend=dict(orientation="v", yanchor="middle", y=0.5, xanchor="left", x=1.05),
            paper_bgcolor='white',
            font=dict(size=11)
        )

        return fig


    # Callback for table content
    @dash_app.callback(
        Output('table-content', 'children'),
        [Input('table-tabs', 'active_tab'),
         Input('session-data', 'data'),
         Input('data-ready-signal', 'children')],
        prevent_initial_call=True
    )
    def render_table_content(active_tab, session_data, ready_signal):
        if not session_data or not ready_signal:
            return html.Div("Upload a file to view summary tables", className='text-muted text-center p-4')

        filepath = session_data.get('filepath')
        processor = get_processor(filepath)
        if not processor:
            return html.Div("Loading data...", className='text-muted text-center p-4')

        tables = processor.get_all_tables()

        table_map = {
            'tab-1': 'stage_footage_summary',
            'tab-2': 'stage_by_pipe_size',
            'tab-3': 'pipe_size_mix',
            'tab-4': 'length_bins',
            'tab-5': 'easement_traffic_summary'
        }

        table_data = tables[table_map[active_tab]]
        df = pd.DataFrame(table_data)

        # Format percentage columns
        for col in df.columns:
            if 'Pct' in col or '%' in col:
                df[col] = df[col].apply(lambda x: f"{x*100:.2f}%" if pd.notna(x) else "")
            elif df[col].dtype in ['float64', 'int64'] and col not in ['Pipe Size', 'Min_Length_ft', 'Max_Length_ft']:
                df[col] = df[col].apply(lambda x: f"{x:,.2f}" if pd.notna(x) else "")

        return dash_table.DataTable(
            data=df.to_dict('records'),
            columns=[{'name': i, 'id': i} for i in df.columns],
            style_cell={
                'textAlign': 'left',
                'padding': '10px',
                'fontFamily': 'Arial'
            },
            style_header={
                'backgroundColor': 'rgb(68, 114, 196)',
                'color': 'white',
                'fontWeight': 'bold'
            },
            style_data_conditional=[
                {
                    'if': {'row_index': 'odd'},
                    'backgroundColor': 'rgb(248, 248, 248)'
                }
            ]
        )


    # Callback for original Excel table
    @dash_app.callback(
        Output('excel-table-container', 'children'),
        [Input('session-data', 'data'),
         Input('data-ready-signal', 'children')],
        prevent_initial_call=True
    )
    def render_excel_table(session_data, ready_signal):
        if not session_data or not ready_signal:
            return html.Div("Upload a file to view original Excel data", className='text-muted text-center p-4')

        filepath = session_data.get('filepath')
        if not filepath or not os.path.exists(filepath):
            return html.Div("Loading data...", className='text-muted text-center p-4')

        # Read the original Excel file
        from openpyxl import load_workbook
        wb = load_workbook(filepath, data_only=True)

        # Find the main sheet
        sheet_name = None
        for name in wb.sheetnames:
            if 'MOINES' in name.upper() or 'SHOT' in name.upper():
                sheet_name = name
                break

        if not sheet_name:
            sheet_name = wb.sheetnames[0]

        ws = wb[sheet_name]

        # Convert to pandas DataFrame
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)

        df = pd.DataFrame(data[1:], columns=data[0])

        # Clean up
        df = df.fillna('')

        # Format date columns nicely
        for col in df.columns:
            if 'Date' in col or 'date' in col:
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d').fillna('')

        # Define intelligent column widths based on content type
        column_widths = {}
        for col in df.columns:
            col_str = str(col)
            # ID columns - narrow
            if 'ID' in col_str.upper() or col_str == 'VIDEO ID':
                column_widths[col_str] = '80px'
            # Segment names - medium
            elif 'SEGMENT' in col_str.upper():
                column_widths[col_str] = '150px'
            # Pipe size - narrow
            elif 'PIPE SIZE' in col_str.upper() or 'SIZE' in col_str.upper():
                column_widths[col_str] = '90px'
            # Length/footage - narrow-medium
            elif 'LENGTH' in col_str.upper() or 'FOOTAGE' in col_str.upper():
                column_widths[col_str] = '100px'
            # Boolean columns (TRUE/FALSE) - narrow
            elif 'COMPLETE' in col_str.upper() or 'READY' in col_str.upper() or 'EASEMENT' in col_str.upper() or 'TRAFFIC' in col_str.upper():
                column_widths[col_str] = '90px'
            # Date columns - medium
            elif 'DATE' in col_str.upper():
                column_widths[col_str] = '110px'
            # Long text columns - wider
            elif 'NOTES' in col_str.upper() or 'COMMENTS' in col_str.upper() or 'DESCRIPTION' in col_str.upper():
                column_widths[col_str] = '250px'
            # Default - medium
            else:
                column_widths[col_str] = '120px'

        # Build style_cell_conditional for column widths
        style_cell_conditional = [
            {
                'if': {'column_id': col},
                'width': width,
                'minWidth': width,
                'maxWidth': width,
                'overflow': 'hidden',
                'textOverflow': 'ellipsis',
            }
            for col, width in column_widths.items()
        ]

        # Add text alignment based on content
        for col in df.columns:
            col_str = str(col)
            if 'ID' in col_str.upper() or 'SIZE' in col_str.upper() or 'LENGTH' in col_str.upper() or 'FOOTAGE' in col_str.upper():
                style_cell_conditional.append({
                    'if': {'column_id': col_str},
                    'textAlign': 'center'
                })

        # Conditional formatting for data rows
        style_data_conditional = [
            # Zebra striping
            {
                'if': {'row_index': 'odd'},
                'backgroundColor': 'rgb(248, 249, 250)'
            }
        ]

        # Add highlighting for TRUE values in completion columns
        for col in df.columns:
            if 'COMPLETE' in str(col).upper() or 'READY' in str(col).upper():
                style_data_conditional.append({
                    'if': {
                        'filter_query': '{{{col}}} = "TRUE"'.format(col=col),
                        'column_id': col
                    },
                    'backgroundColor': '#d4edda',
                    'color': '#155724',
                    'fontWeight': 'bold'
                })

        return dash_table.DataTable(
            data=df.to_dict('records'),
            columns=[{'name': str(i), 'id': str(i)} for i in df.columns],
            style_cell={
                'textAlign': 'left',
                'padding': '10px 12px',
                'fontSize': '12px',
                'fontFamily': 'Arial, sans-serif',
                'whiteSpace': 'normal',
                'height': 'auto',
            },
            style_cell_conditional=style_cell_conditional,
            style_header={
                'backgroundColor': '#4472C4',
                'color': 'white',
                'fontWeight': 'bold',
                'textAlign': 'center',
                'padding': '12px',
                'fontSize': '13px',
                'border': '1px solid #2c5aa0',
                'whiteSpace': 'normal',
                'height': 'auto',
            },
            style_data={
                'border': '1px solid #dee2e6',
                'whiteSpace': 'normal',
                'height': 'auto',
            },
            style_data_conditional=style_data_conditional,
            style_table={
                'overflowX': 'auto',
                'maxHeight': '650px',
                'overflowY': 'auto',
                'border': '2px solid #4472C4',
                'boxShadow': '0 2px 8px rgba(0,0,0,0.1)'
            },
            fixed_rows={'headers': True},
            page_size=50,
            tooltip_data=[
                {
                    column: {'value': str(value), 'type': 'markdown'}
                    for column, value in row.items()
                } for row in df.to_dict('records')
            ],
            tooltip_duration=None,
            css=[{
                'selector': '.dash-table-tooltip',
                'rule': 'background-color: #333; color: white; padding: 8px; border-radius: 4px; font-size: 12px;'
            }]
        )


    # Callback for downloads
    @dash_app.callback(
        Output('download-file', 'data'),
        [Input('download-btn-1', 'n_clicks'),
         Input('download-btn-2', 'n_clicks'),
         Input('download-btn-3', 'n_clicks')],
        [State('session-data', 'data')],
        prevent_initial_call=True
    )
    def download_excel(btn1, btn2, btn3, session_data):
        if not session_data:
            raise PreventUpdate

        ctx = dash.callback_context
        if not ctx.triggered:
            raise PreventUpdate

        button_id = ctx.triggered[0]['prop_id'].split('.')[0]

        approach_map = {
            'download-btn-1': 'approach1',
            'download-btn-2': 'approach2',
            'download-btn-3': 'approach3'
        }

        approach = approach_map.get(button_id)
        if not approach:
            raise PreventUpdate

        session_id = session_data['session_id']
        filepath = session_data.get('filepath')
        processor = get_processor(filepath)

        if not processor or not filepath:
            raise PreventUpdate

        # Generate Excel file (modifying original)
        generator = ExcelDashboardGeneratorV2(processor, filepath)

        output_path = outputs_dir / f"{session_id}_{approach}.xlsx"

        if approach == 'approach1':
            generator.generate_approach_1(str(output_path))
        elif approach == 'approach2':
            generator.generate_approach_2(str(output_path))
        elif approach == 'approach3':
            generator.generate_approach_3(str(output_path))

        return dcc.send_file(str(output_path))


    # Callback to toggle prep complete state
    @dash_app.callback(
        Output('prep-toggle-state', 'data'),
        Input('kpi-prep-complete-card', 'n_clicks'),
        State('prep-toggle-state', 'data'),
        prevent_initial_call=True
    )
    def toggle_prep_state(n_clicks, current_state):
        if n_clicks is None:
            raise PreventUpdate
        return {'show_fraction': not current_state['show_fraction']}


    # Callback to toggle completion state
    @dash_app.callback(
        Output('completion-toggle-state', 'data'),
        Input('kpi-completion-card', 'n_clicks'),
        State('completion-toggle-state', 'data'),
        prevent_initial_call=True
    )
    def toggle_completion_state(n_clicks, current_state):
        if n_clicks is None:
            raise PreventUpdate
        return {'show_fraction': not current_state['show_fraction']}


    # Callback to update prep complete display
    @dash_app.callback(
        Output('kpi-prep-complete', 'children'),
        [Input('prep-toggle-state', 'data'),
         Input('session-data', 'data'),
         Input('data-ready-signal', 'children')],
        prevent_initial_call=True
    )
    def update_prep_complete_display(toggle_state, session_data, ready_signal):
        if not session_data or not ready_signal:
            raise PreventUpdate

        filepath = session_data.get('filepath')
        processor = get_processor(filepath)
        if not processor:
            raise PreventUpdate

        tables = processor.get_all_tables()
        stage_summary = tables['stage_footage_summary']

        ready_to_line_count = sum(
            r['Segment_Count'] for r in stage_summary
            if r['Stage'] == 'Ready to Line'
        )
        total_segments = len(processor.segments)

        if toggle_state.get('show_fraction', False):
            return f"{ready_to_line_count}/{total_segments}"
        else:
            prep_complete_pct = (ready_to_line_count / total_segments * 100) if total_segments > 0 else 0
            return f"{prep_complete_pct:.1f}%"


    # Callback to update completion display
    @dash_app.callback(
        Output('kpi-completion', 'children'),
        [Input('completion-toggle-state', 'data'),
         Input('session-data', 'data'),
         Input('data-ready-signal', 'children')],
        prevent_initial_call=True
    )
    def update_completion_display(toggle_state, session_data, ready_signal):
        if not session_data or not ready_signal:
            raise PreventUpdate

        filepath = session_data.get('filepath')
        processor = get_processor(filepath)
        if not processor:
            raise PreventUpdate

        tables = processor.get_all_tables()
        stage_summary = tables['stage_footage_summary']

        completed_count = sum(
            r['Segment_Count'] for r in stage_summary
            if r['Stage'] in ['Lined', 'Post TV Complete', 'Grouted/Done']
        )
        total_segments = len(processor.segments)

        if toggle_state.get('show_fraction', False):
            return f"{completed_count}/{total_segments}"
        else:
            completed_footage = sum(
                r['Total_Feet'] for r in stage_summary
                if r['Stage'] in ['Lined', 'Post TV Complete', 'Grouted/Done']
            )
            completion_pct = (completed_footage / processor.total_footage * 100) if processor.total_footage > 0 else 0
            return f"{completion_pct:.1f}%"

    # ============================================================================
    # BREAKOUT TABLES - Interactive Filtered Segment Views
    # ============================================================================

    @dash_app.callback(
        Output('breakout-table-content', 'children'),
        [Input('breakout-tabs', 'active_tab'),
         Input('breakout-nav-store', 'data'),
         Input('session-data', 'data')]
    )
    def update_breakout_table(active_tab, nav_data, session_data):
        """
        Display filtered breakout tables based on tab selection or chart clicks.

        Tables show segments matching specific business criteria.
        """
        if not session_data:
            return html.Div([
                html.I(className="fas fa-info-circle fa-3x text-muted mb-3"),
                html.P("Upload a file to see breakout tables", className='text-muted text-center')
            ], className='text-center p-5')

        processor = get_processor(session_data['filepath'])
        if not processor:
            return html.P("Error loading data", className='text-danger text-center p-5')

        # Determine filter based on tab or navigation click
        filter_type = active_tab
        filter_value = None

        if nav_data:
            filter_type = nav_data.get('type', active_tab)
            filter_value = nav_data.get('value')

        # Get filtered segments based on filter type
        segments = []
        title = ""
        subtitle = ""

        if filter_type == 'breakout-awaiting':
            segments = processor.get_segments_awaiting_prep()
            title = f"Awaiting Prep Segments"
            subtitle = f"{len(segments)} segments with map length but no prep crew verified diameter"

        elif filter_type == 'breakout-ready':
            segments = processor.get_segments_ready_to_line()
            title = f"Ready to Line Segments"
            subtitle = f"{len(segments)} segments certified ready but not yet lined"

        elif filter_type == 'breakout-cctv':
            segments = processor.get_segments_cctv_posted()
            title = f"CCTV Posted (Post TV Complete) Segments"
            subtitle = f"{len(segments)} segments with final post-TV inspection complete"

        elif filter_type == 'breakout-pending':
            segments = processor.get_segments_pending()
            title = f"Pending Segments"
            subtitle = f"{len(segments)} segments with prep data but not yet certified ready"

        elif filter_type == 'breakout-easement':
            segments = processor.get_segments_by_easement(True)
            title = f"Easement Segments"
            subtitle = f"{len(segments)} segments requiring easement access"

        elif filter_type == 'breakout-traffic':
            segments = processor.get_segments_by_traffic_control(True)
            title = f"Traffic Control Required"
            subtitle = f"{len(segments)} segments requiring traffic control"

        elif filter_type == 'breakout-row':
            segments = processor.get_segments_row_only()
            title = f"ROW Only Segments"
            subtitle = f"{len(segments)} segments with no easement or traffic control requirements"

        elif filter_type == 'breakout-pipe':
            if filter_value:
                segments = processor.get_segments_by_pipe_size(filter_value)
                title = f"{filter_value}\" Pipe Segments"
                subtitle = f"{len(segments)} segments with {filter_value}\" diameter pipe"
            else:
                segments = []
                title = "Pipe Size Filter"
                subtitle = "Click on a pipe size in the charts to filter"

        elif filter_type == 'breakout-length':
            if filter_value and 'min' in filter_value:
                segments = processor.get_segments_by_length_bin(
                    filter_value['min'],
                    filter_value.get('max')
                )
                bin_label = filter_value.get('label', 'Selected Range')
                title = f"{bin_label} Length Segments"
                subtitle = f"{len(segments)} segments in this length range"
            else:
                segments = []
                title = "Length Range Filter"
                subtitle = "Click on a length bin in the chart to filter"

        else:  # breakout-all
            segments = processor.segments
            title = f"All Segments"
            subtitle = f"{len(segments)} total segments in project"

        # Handle empty results
        if not segments:
            return html.Div([
                html.H6(title, className='mb-2'),
                html.P(subtitle, className='text-muted mb-4'),
                html.Div([
                    html.I(className="fas fa-inbox fa-3x text-muted mb-3"),
                    html.P("No segments match this filter", className='text-muted')
                ], className='text-center p-5')
            ])

        # Format for table display
        formatted = processor.format_segments_for_table(segments)
        df = pd.DataFrame(formatted)

        # Calculate statistics
        total_footage = sum(s['map_length'] for s in segments)
        avg_length = total_footage / len(segments) if segments else 0
        pct_of_total = (total_footage / processor.total_footage * 100) if processor.total_footage > 0 else 0

        # Create statistics footer
        stats_footer = dbc.Row([
            dbc.Col([
                html.Div([
                    html.I(className="fas fa-list text-primary me-2"),
                    html.Strong("Segments: "),
                    html.Span(f"{len(segments)}")
                ])
            ], width=3),
            dbc.Col([
                html.Div([
                    html.I(className="fas fa-ruler text-success me-2"),
                    html.Strong("Total Footage: "),
                    html.Span(f"{total_footage:,.0f} ft ({pct_of_total:.1f}% of project)")
                ])
            ], width=4),
            dbc.Col([
                html.Div([
                    html.I(className="fas fa-calculator text-info me-2"),
                    html.Strong("Avg Length: "),
                    html.Span(f"{avg_length:.1f} ft")
                ])
            ], width=3),
            dbc.Col([
                dbc.Button([
                    html.I(className="fas fa-file-export me-2"),
                    "Export to Excel"
                ], color="primary", size="sm", outline=True)
            ], width=2, className='text-end')
        ], className='mt-4 p-3 bg-light rounded align-items-center')

        return html.Div([
            html.H5(title, className='mb-1'),
            html.P(subtitle, className='text-muted mb-3'),
            dash_table.DataTable(
                data=df.to_dict('records'),
                columns=[{"name": i, "id": i} for i in df.columns],
                style_table={'overflowX': 'auto'},
                style_cell={
                    'textAlign': 'left',
                    'padding': '10px',
                    'minWidth': '100px',
                    'fontSize': '14px'
                },
                style_header={
                    'backgroundColor': '#1E3A8A',
                    'color': 'white',
                    'fontWeight': 'bold',
                    'textAlign': 'left'
                },
                style_data_conditional=[
                    {
                        'if': {'row_index': 'odd'},
                        'backgroundColor': 'rgb(248, 248, 248)'
                    },
                    {
                        'if': {'column_id': 'Flagged Issues', 'filter_query': '{Flagged Issues} ne "-"'},
                        'backgroundColor': 'rgba(239, 68, 68, 0.1)',
                        'color': '#991b1b'
                    }
                ],
                page_size=25,
                page_action='native',
                sort_action='native',
                filter_action='native',
                export_format='xlsx',
                export_headers='display'
            ),
            stats_footer
        ])

    # ============================================================================
    # CHART CLICK HANDLERS - Navigate to Breakout Tables
    # ============================================================================

    @dash_app.callback(
        [Output('breakout-nav-store', 'data'),
         Output('breakout-tabs', 'active_tab')],
        [Input('progress-bar-chart', 'clickData'),
         Input('pipe-progress-chart', 'clickData'),
         Input('pipe-size-chart', 'clickData'),
         Input('easement-traffic-chart', 'clickData'),
         Input('length-distribution-chart', 'clickData')]
    )
    def handle_chart_clicks(progress_click, pipe_progress_click, pipe_size_click,
                           easement_click, length_click):
        """
        Handle clicks on charts to navigate to appropriate breakout table.

        When a user clicks on a chart element, this automatically switches
        to the relevant breakout tab and filters the segments.
        """
        ctx = dash.callback_context

        if not ctx.triggered:
            raise PreventUpdate

        trigger_id = ctx.triggered[0]['prop_id'].split('.')[0]

        # Progress bar chart click - navigate to that stage
        if trigger_id == 'progress-bar-chart' and progress_click:
            stage = progress_click['points'][0].get('label', progress_click['points'][0].get('y', ''))

            # Map stage to breakout tab
            if 'Ready to Line' in stage:
                return {'type': 'breakout-ready', 'value': stage}, 'breakout-ready'
            elif 'Post TV' in stage or 'CCTV' in stage:
                return {'type': 'breakout-cctv', 'value': stage}, 'breakout-cctv'
            else:
                return {'type': 'breakout-current', 'value': stage}, 'breakout-current'

        # Pipe progress or size chart click - navigate to that pipe size
        elif (trigger_id == 'pipe-progress-chart' or trigger_id == 'pipe-size-chart') and (pipe_progress_click or pipe_size_click):
            click_data = pipe_progress_click if pipe_progress_click else pipe_size_click
            pipe_size = click_data['points'][0].get('y', click_data['points'][0].get('x', ''))
            return {'type': 'breakout-pipe', 'value': pipe_size}, 'breakout-current'

        # Easement/traffic chart click
        elif trigger_id == 'easement-traffic-chart' and easement_click:
            label = easement_click['points'][0].get('x', '')

            if 'Easement' in label and 'Yes' in label:
                return {'type': 'breakout-easement', 'value': True}, 'breakout-easement'
            elif 'Traffic' in label and 'Yes' in label:
                return {'type': 'breakout-traffic', 'value': True}, 'breakout-traffic'
            else:
                return {}, 'breakout-all'

        # Length distribution chart click
        elif trigger_id == 'length-distribution-chart' and length_click:
            bin_label = length_click['points'][0].get('x', '')

            # Parse bin label to extract min/max
            # This is a simplified version - you may need to enhance based on actual bin labels
            return {'type': 'breakout-length', 'value': {'label': bin_label}}, 'breakout-current'

        raise PreventUpdate

    return dash_app
