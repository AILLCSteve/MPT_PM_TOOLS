"""
Core data processing module for CIPP shot schedule analysis.
Handles reading Excel data, computing stages, and building summary tables.
"""

from openpyxl import load_workbook
from typing import List, Dict, Any, Optional
from collections import defaultdict


class CIPPDataProcessor:
    """Process CIPP shot schedule data and generate summary tables."""

    STAGES = [
        "Not Started",
        "Awaiting Prep",
        "Ready to Line",
        "Wet Out",
        "Lined",
        "Post TV Complete"
    ]

    LENGTH_BINS = [
        {"label": "0–50", "min": 0, "max": 50},
        {"label": "51–150", "min": 51, "max": 150},
        {"label": "151–250", "min": 151, "max": 250},
        {"label": "251–350", "min": 251, "max": 350},
        {"label": "351+", "min": 351, "max": None},
    ]

    def __init__(self, file_path: str, sheet_name: str = "WEST DES MOINES, IA Shot Schedu"):
        """Initialize processor with Excel file path."""
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.segments = []
        self.total_footage = 0

    def load_data(self) -> List[Dict[str, Any]]:
        """Load and validate data from Excel file."""
        wb = load_workbook(self.file_path, data_only=True)

        # Try to find the sheet
        if self.sheet_name not in wb.sheetnames:
            # Try to find sheet with similar name
            for sheet in wb.sheetnames:
                if "MOINES" in sheet.upper() or "SHOT" in sheet.upper():
                    self.sheet_name = sheet
                    break
            else:
                raise ValueError(f"Sheet '{self.sheet_name}' not found. Available: {wb.sheetnames}")

        ws = wb[self.sheet_name]

        # Get headers from row 1
        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value:
                headers[str(cell.value).strip()] = col_idx

        # Process data rows
        segments = []
        for row_idx in range(2, ws.max_row + 1):
            row = ws[row_idx]

            # Get VIDEO ID and Map Length
            video_id = self._get_cell_value(row, headers.get("VIDEO ID"))
            map_length = self._get_cell_value(row, headers.get("Map Length"))

            # Validate segment
            if not self._is_valid_segment(video_id, map_length):
                continue

            # Extract segment data
            segment = {
                "video_id": video_id,
                "line_segment": self._get_cell_value(row, headers.get("Line Segment")),
                "pipe_size": self._get_cell_value(row, headers.get("Pipe Size")),
                "map_length": float(map_length),
                "prep_complete": self._is_truthy(self._get_cell_value(row, headers.get("Prep Complete"))),
                "prep_crew_verified_dia": self._get_cell_value(row, headers.get("Prep Crew Verified Dia")),
                "prep_usmh_depth": self._get_cell_value(row, headers.get("Prep USMH Depth")),
                "prep_dsmh_depth": self._get_cell_value(row, headers.get("Prep DSMH Depth")),
                "ready_to_line": self._is_truthy(self._get_cell_value(row, headers.get("Ready to Line - Certified by Prep Crew Lead"))),
                "wet_out_date": self._get_cell_value(row, headers.get("Wet Out Date")),
                "lining_date": self._get_cell_value(row, headers.get("Lining Date")),
                "final_post_tv_date": self._get_cell_value(row, headers.get("Final Post TV Date")),
                "grout_state_date": self._get_cell_value(row, headers.get("Grout State Date")),
                "easement": self._is_truthy(self._get_cell_value(row, headers.get("Easement"))),
                "traffic_control": self._is_truthy(self._get_cell_value(row, headers.get("Traffic Control"))),
            }

            # Compute current stage (for display/sorting)
            segment["stage"] = self._compute_stage(segment)

            # Compute ALL achieved stages (CUMULATIVE LIFECYCLE)
            segment["achieved_stages"] = self._compute_achieved_stages(segment)

            segments.append(segment)

        self.segments = segments
        self.total_footage = sum(s["map_length"] for s in segments)
        return segments

    def _get_cell_value(self, row, col_idx):
        """Get cell value safely."""
        if col_idx is None or col_idx < 1:
            return None
        try:
            return row[col_idx - 1].value
        except (IndexError, AttributeError):
            return None

    def _is_valid_segment(self, video_id, map_length) -> bool:
        """Check if segment is valid."""
        try:
            video_id_num = float(video_id) if video_id is not None else 0
            map_length_num = float(map_length) if map_length is not None else 0
            return video_id_num >= 1 and map_length_num > 0
        except (ValueError, TypeError):
            return False

    def _is_truthy(self, value) -> bool:
        """Check if value is truthy."""
        if value is None:
            return False
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return value != 0
        if isinstance(value, str):
            return value.upper() in ["TRUE", "YES", "Y", "1"]
        return bool(value)

    def _compute_stage(self, segment: Dict[str, Any]) -> str:
        """
        Compute current stage based on priority logic (for display).

        This is MUTUALLY EXCLUSIVE - each segment is in ONE stage at a time.
        Used for the overall progress bar to avoid overlapping areas.
        """
        if segment["final_post_tv_date"] is not None:
            return "Post TV Complete"
        elif segment["lining_date"] is not None:
            return "Lined"
        elif segment["wet_out_date"] is not None:
            return "Wet Out"
        elif segment["ready_to_line"]:
            return "Ready to Line"
        elif segment["prep_complete"]:
            return "Awaiting Prep"
        else:
            return "Not Started"

    def _compute_achieved_stages(self, segment: Dict[str, Any]) -> List[str]:
        """
        Compute ALL stages this segment has achieved (CUMULATIVE LIFECYCLE).

        This is the transformative change: segments don't leave previous stages
        when they advance. A "Lined" segment HAS ALSO achieved "Awaiting Prep",
        "Ready to Line", and "Wet Out". This gives accurate progress tracking.

        Returns list of all achieved stage names.
        """
        achieved = []

        # Check each milestone in order (cumulative progression)
        if segment["prep_complete"] or segment["ready_to_line"] or segment["wet_out_date"] or segment["lining_date"] or segment["final_post_tv_date"]:
            achieved.append("Awaiting Prep")

        if segment["ready_to_line"] or segment["wet_out_date"] or segment["lining_date"] or segment["final_post_tv_date"]:
            achieved.append("Ready to Line")

        if segment["wet_out_date"] or segment["lining_date"] or segment["final_post_tv_date"]:
            achieved.append("Wet Out")

        if segment["lining_date"] or segment["final_post_tv_date"]:
            achieved.append("Lined")

        if segment["final_post_tv_date"]:
            achieved.append("Post TV Complete")

        # If nothing achieved, it's not started
        if not achieved:
            achieved.append("Not Started")

        return achieved

    def get_overall_progress_summary(self) -> List[Dict[str, Any]]:
        """
        Overall Progress Summary (MUTUALLY EXCLUSIVE for progress bar).

        This uses the OLD counting method where each segment is in ONE stage.
        Used specifically for the overall progress bar to avoid overlapping areas.

        Returns stage summary with mutually exclusive counts.
        """
        stage_data = defaultdict(lambda: {"count": 0, "footage": 0})

        for segment in self.segments:
            # Count this segment ONLY in its current stage (mutually exclusive)
            stage = segment["stage"]
            stage_data[stage]["count"] += 1
            stage_data[stage]["footage"] += segment["map_length"]

        result = []
        for stage in self.STAGES:
            data = stage_data[stage]
            result.append({
                "Stage": stage,
                "Segment_Count": data["count"],
                "Total_Feet": round(data["footage"], 2),
                "Pct_of_Total_Feet": round(data["footage"] / self.total_footage, 4) if self.total_footage > 0 else 0
            })

        return result

    def get_stage_footage_summary(self) -> List[Dict[str, Any]]:
        """
        Table 1: Stage_Footage_Summary (CUMULATIVE LIFECYCLE tracking).

        IMPORTANT: This uses cumulative counting where segments are counted
        for EVERY stage they've achieved, not just their current stage.

        A segment that's "Lined" is counted in:
        - Awaiting Prep (it achieved that)
        - Ready to Line (it achieved that)
        - Wet Out (it achieved that)
        - Lined (it achieved that)

        This gives accurate progress metrics showing how much has passed
        through each lifecycle gate.
        """
        stage_data = defaultdict(lambda: {"count": 0, "footage": 0})

        for segment in self.segments:
            # Count this segment for EVERY stage it has achieved
            for achieved_stage in segment["achieved_stages"]:
                stage_data[achieved_stage]["count"] += 1
                stage_data[achieved_stage]["footage"] += segment["map_length"]

        result = []
        for stage in self.STAGES:
            data = stage_data[stage]
            result.append({
                "Stage": stage,
                "Segment_Count": data["count"],
                "Total_Feet": round(data["footage"], 2),
                "Pct_of_Total_Feet": round(data["footage"] / self.total_footage, 4) if self.total_footage > 0 else 0
            })

        return result

    def get_stage_by_pipe_size(self) -> List[Dict[str, Any]]:
        """
        Table 2: Stage_by_PipeSize (CUMULATIVE progress by diameter).

        Uses cumulative counting: a segment is counted for every stage
        it has achieved, not just its current stage.
        """
        # Collect unique pipe sizes
        pipe_sizes = sorted(set(s["pipe_size"] for s in self.segments if s["pipe_size"] is not None))

        result = []
        for pipe_size in pipe_sizes:
            row = {"Pipe Size": pipe_size}

            # For each stage, count all segments that have ACHIEVED it
            for stage in self.STAGES:
                footage = sum(
                    s["map_length"]
                    for s in self.segments
                    if s["pipe_size"] == pipe_size and stage in s["achieved_stages"]
                )
                row[stage] = round(footage, 2)

            # Total is just the total footage of this pipe size (not cumulative)
            row["Total_Feet"] = round(sum(
                s["map_length"]
                for s in self.segments
                if s["pipe_size"] == pipe_size
            ), 2)
            result.append(row)

        return result

    def get_pipe_size_mix(self) -> List[Dict[str, Any]]:
        """Table 3: Pipe_Size_Mix (how much of each diameter)."""
        pipe_sizes = sorted(set(s["pipe_size"] for s in self.segments if s["pipe_size"] is not None))

        result = []
        for pipe_size in pipe_sizes:
            segments_for_size = [s for s in self.segments if s["pipe_size"] == pipe_size]
            count = len(segments_for_size)
            footage = sum(s["map_length"] for s in segments_for_size)

            result.append({
                "Pipe Size": pipe_size,
                "Segment_Count": count,
                "Total_Feet": round(footage, 2),
                "Avg_Length_ft": round(footage / count, 2) if count > 0 else 0,
                "Pct_of_Total_Feet": round(footage / self.total_footage, 4) if self.total_footage > 0 else 0
            })

        return result

    def get_length_bins(self) -> List[Dict[str, Any]]:
        """Table 4: Length_Bins (distribution of run lengths)."""
        result = []

        for bin_def in self.LENGTH_BINS:
            segments_in_bin = []
            for segment in self.segments:
                length = segment["map_length"]
                min_len = bin_def["min"]
                max_len = bin_def["max"]

                if max_len is None:
                    # Open-ended bin
                    if length >= min_len:
                        segments_in_bin.append(segment)
                else:
                    if min_len <= length <= max_len:
                        segments_in_bin.append(segment)

            count = len(segments_in_bin)
            footage = sum(s["map_length"] for s in segments_in_bin)

            result.append({
                "Length_Bin_Label": bin_def["label"],
                "Min_Length_ft": bin_def["min"],
                "Max_Length_ft": bin_def["max"] if bin_def["max"] is not None else "",
                "Segment_Count": count,
                "Total_Feet": round(footage, 2),
                "Pct_of_Total_Feet": round(footage / self.total_footage, 4) if self.total_footage > 0 else 0
            })

        return result

    def get_easement_traffic_summary(self) -> List[Dict[str, Any]]:
        """Table 5: Easement_Traffic_Summary."""
        result = []

        # Easement Yes
        easement_yes = [s for s in self.segments if s["easement"]]
        result.append({
            "Category": "Easement",
            "Flag": "Yes",
            "Segment_Count": len(easement_yes),
            "Total_Feet": round(sum(s["map_length"] for s in easement_yes), 2),
            "Pct_of_Total_Feet": round(sum(s["map_length"] for s in easement_yes) / self.total_footage, 4) if self.total_footage > 0 else 0
        })

        # Easement No
        easement_no = [s for s in self.segments if not s["easement"]]
        result.append({
            "Category": "Easement",
            "Flag": "No",
            "Segment_Count": len(easement_no),
            "Total_Feet": round(sum(s["map_length"] for s in easement_no), 2),
            "Pct_of_Total_Feet": round(sum(s["map_length"] for s in easement_no) / self.total_footage, 4) if self.total_footage > 0 else 0
        })

        # Traffic Control Yes
        traffic_yes = [s for s in self.segments if s["traffic_control"]]
        result.append({
            "Category": "Traffic Control",
            "Flag": "Yes",
            "Segment_Count": len(traffic_yes),
            "Total_Feet": round(sum(s["map_length"] for s in traffic_yes), 2),
            "Pct_of_Total_Feet": round(sum(s["map_length"] for s in traffic_yes) / self.total_footage, 4) if self.total_footage > 0 else 0
        })

        # Traffic Control No
        traffic_no = [s for s in self.segments if not s["traffic_control"]]
        result.append({
            "Category": "Traffic Control",
            "Flag": "No",
            "Segment_Count": len(traffic_no),
            "Total_Feet": round(sum(s["map_length"] for s in traffic_no), 2),
            "Pct_of_Total_Feet": round(sum(s["map_length"] for s in traffic_no) / self.total_footage, 4) if self.total_footage > 0 else 0
        })

        return result

    def get_all_tables(self) -> Dict[str, List[Dict[str, Any]]]:
        """Get all 5 summary tables."""
        return {
            "stage_footage_summary": self.get_stage_footage_summary(),
            "stage_by_pipe_size": self.get_stage_by_pipe_size(),
            "pipe_size_mix": self.get_pipe_size_mix(),
            "length_bins": self.get_length_bins(),
            "easement_traffic_summary": self.get_easement_traffic_summary()
        }

    # ========================================================================
    # BREAKOUT TABLE FILTERING FUNCTIONS
    # ========================================================================

    def get_segments_by_achieved_stage(self, stage: str) -> List[Dict[str, Any]]:
        """
        Get all segments that have ACHIEVED a specific stage.

        This is the power of cumulative tracking: you can see ALL segments
        that have passed through a specific lifecycle gate.

        Args:
            stage: Stage name (e.g., "Ready to Line", "Prep Complete")

        Returns:
            List of segment dictionaries that have achieved this stage
        """
        return [s for s in self.segments if stage in s["achieved_stages"]]

    def get_segments_by_current_stage(self, stage: str) -> List[Dict[str, Any]]:
        """
        Get all segments currently AT a specific stage (not yet advanced further).

        Args:
            stage: Stage name

        Returns:
            List of segment dictionaries currently at this stage
        """
        return [s for s in self.segments if s["stage"] == stage]

    def get_segments_by_pipe_size(self, pipe_size: float) -> List[Dict[str, Any]]:
        """Filter segments by pipe size."""
        return [s for s in self.segments if s["pipe_size"] == pipe_size]

    def get_segments_by_length_bin(self, min_length: float, max_length: Optional[float] = None) -> List[Dict[str, Any]]:
        """Filter segments by length range."""
        if max_length is None:
            return [s for s in self.segments if s["map_length"] >= min_length]
        return [s for s in self.segments if min_length <= s["map_length"] <= max_length]

    def get_segments_by_easement(self, is_easement: bool = True) -> List[Dict[str, Any]]:
        """Filter segments by easement status."""
        return [s for s in self.segments if s["easement"] == is_easement]

    def get_segments_by_traffic_control(self, requires_traffic: bool = True) -> List[Dict[str, Any]]:
        """Filter segments by traffic control requirement."""
        return [s for s in self.segments if s["traffic_control"] == requires_traffic]

    def get_segments_flagged_for_issues(self) -> List[Dict[str, Any]]:
        """
        Get segments flagged for potential issues.

        Currently flags segments that:
        - Are in easements
        - Require traffic control
        - Are not started but should be (heuristic: project > 30 days old)

        Can be extended with more heuristics.
        """
        flagged = []
        for s in self.segments:
            issues = []
            if s["easement"]:
                issues.append("Easement Access")
            if s["traffic_control"]:
                issues.append("Traffic Control")
            if s["stage"] == "Not Started":
                issues.append("Not Yet Started")

            if issues:
                segment_copy = s.copy()
                segment_copy["flagged_issues"] = issues
                flagged.append(segment_copy)

        return flagged

    def get_segments_ready_to_line(self) -> List[Dict[str, Any]]:
        """Get segments with ready_to_line=true but not yet lined (no lining date)."""
        return [s for s in self.segments if s["ready_to_line"] and not s["lining_date"]]

    def get_segments_cctv_posted(self) -> List[Dict[str, Any]]:
        """Get segments with Post TV complete (final_post_tv_date is not null)."""
        return [s for s in self.segments if s["final_post_tv_date"] is not None]

    def get_segments_pending(self) -> List[Dict[str, Any]]:
        """Get segments with prep data but not ready to line (Pending status)."""
        return [
            s for s in self.segments
            if (s.get("prep_usmh_depth") or s.get("prep_dsmh_depth") or s.get("prep_crew_verified_dia"))
            and not s["ready_to_line"]
        ]

    def get_segments_row_only(self) -> List[Dict[str, Any]]:
        """Get segments where both easement and traffic control are false (ROW only)."""
        return [s for s in self.segments if not s["easement"] and not s["traffic_control"]]

    def get_segments_awaiting_prep(self) -> List[Dict[str, Any]]:
        """Get segments with map length but no prep crew verified diameter (Prep not started)."""
        return [
            s for s in self.segments
            if s["map_length"] > 0 and (s.get("prep_crew_verified_dia") is None or s.get("prep_crew_verified_dia") == "")
        ]

    def format_segments_for_table(self, segments: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Format segments for display in breakout tables.

        Returns list of dicts with display-friendly keys and formatting.
        """
        formatted = []
        for s in segments:
            formatted.append({
                "Video ID": s["video_id"],
                "Line Segment": s["line_segment"] or "-",
                "Pipe Size": s["pipe_size"] or "-",
                "Map Length (ft)": round(s["map_length"], 1),
                "Current Stage": s["stage"],
                "Achieved Stages": ", ".join(s["achieved_stages"]),
                "Easement": "Yes" if s["easement"] else "No",
                "Traffic Control": "Yes" if s["traffic_control"] else "No",
                "Flagged Issues": ", ".join(s.get("flagged_issues", [])) if "flagged_issues" in s else "-"
            })
        return formatted
