"""
Excel Dashboard Generator V2
Modifies the ORIGINAL uploaded Excel file by adding dashboard sheets
"""

from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Dict, List, Any
import io
import shutil


class ExcelDashboardGeneratorV2:
    """Generate Excel dashboards by modifying the original uploaded file."""

    def __init__(self, data_processor, original_filepath):
        """Initialize with processor and path to original file."""
        self.processor = data_processor
        self.original_filepath = original_filepath
        self.tables = data_processor.get_all_tables()

    def generate_approach_1(self, output_path: str):
        """Approach 1: Add openpyxl native charts to original file."""
        # Copy original file
        shutil.copy2(self.original_filepath, output_path)

        # Open the copy
        wb = load_workbook(output_path)

        # Remove old dashboard sheets if they exist
        if "Dashboard_Data" in wb.sheetnames:
            del wb["Dashboard_Data"]
        if "Dashboard" in wb.sheetnames:
            del wb["Dashboard"]

        # Create new sheets
        ws_data = wb.create_sheet("Dashboard_Data", 0)
        ws_dash = wb.create_sheet("Dashboard", 0)

        # Write tables
        self._write_all_tables(ws_data, wb)

        # Create charts
        self._create_openpyxl_charts(ws_dash, ws_data)

        wb.save(output_path)
        return output_path

    def generate_approach_2(self, output_path: str):
        """Approach 2: Use xlsxwriter (requires creating new workbook with original data)."""
        try:
            import xlsxwriter
        except ImportError:
            raise ImportError("xlsxwriter not installed. Run: pip install xlsxwriter")

        # Read original data
        original_wb = load_workbook(self.original_filepath, data_only=True)

        # Create new workbook with xlsxwriter
        wb = xlsxwriter.Workbook(output_path)

        # Copy original sheets first
        for sheet_name in original_wb.sheetnames:
            ws_orig = original_wb[sheet_name]
            ws_new = wb.add_worksheet(sheet_name)

            # Copy data
            for row_idx, row in enumerate(ws_orig.iter_rows(values_only=True)):
                for col_idx, value in enumerate(row):
                    if value is not None:
                        ws_new.write(row_idx, col_idx, value)

        # Create Dashboard_Data sheet
        ws_data = wb.add_worksheet("Dashboard_Data")
        self._write_all_tables_xlsxwriter(wb, ws_data)

        # Create Dashboard sheet with charts
        ws_dash = wb.add_worksheet("Dashboard")
        self._create_xlsxwriter_charts(wb, ws_dash, ws_data)

        wb.close()
        return output_path

    def generate_approach_3(self, output_path: str):
        """Approach 3: Add plotly charts as images to original file."""
        try:
            import plotly.graph_objects as go
        except ImportError:
            raise ImportError("plotly not installed")

        # Copy original file
        shutil.copy2(self.original_filepath, output_path)

        # Open the copy
        wb = load_workbook(output_path)

        # Remove old dashboard sheets if they exist
        if "Dashboard_Data" in wb.sheetnames:
            del wb["Dashboard_Data"]
        if "Dashboard" in wb.sheetnames:
            del wb["Dashboard"]

        # Create new sheets
        ws_data = wb.create_sheet("Dashboard_Data", 0)
        ws_dash = wb.create_sheet("Dashboard", 0)

        # Write tables
        self._write_all_tables(ws_data, wb)

        # Create charts with plotly images
        self._create_plotly_charts(ws_dash, ws_data)

        wb.save(output_path)
        return output_path

    def _write_all_tables(self, ws, wb):
        """Write all 5 tables to Dashboard_Data sheet using openpyxl."""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # Table 1: Stage_Footage_Summary at A1
        self._write_table(ws, 1, 1, self.tables["stage_footage_summary"], header_fill, header_font)

        # Table 2: Stage_by_PipeSize at A10
        self._write_table(ws, 10, 1, self.tables["stage_by_pipe_size"], header_fill, header_font)

        # Table 3: Pipe_Size_Mix at A19
        row_offset = 10 + len(self.tables["stage_by_pipe_size"]) + 2
        self._write_table(ws, row_offset, 1, self.tables["pipe_size_mix"], header_fill, header_font)

        # Table 4: Length_Bins
        row_offset = row_offset + len(self.tables["pipe_size_mix"]) + 2
        self._write_table(ws, row_offset, 1, self.tables["length_bins"], header_fill, header_font)

        # Table 5: Easement_Traffic_Summary
        row_offset = row_offset + len(self.tables["length_bins"]) + 2
        self._write_table(ws, row_offset, 1, self.tables["easement_traffic_summary"], header_fill, header_font)

    def _write_table(self, ws, start_row, start_col, data, header_fill, header_font):
        """Write a single table to the worksheet."""
        if not data:
            return

        # Write headers
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers):
            cell = ws.cell(row=start_row, column=start_col + col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, header in enumerate(headers):
                value = row_data[header]
                cell = ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=value)

                # Format percentages
                if "Pct" in header or "%" in header:
                    cell.number_format = '0.00%'
                elif isinstance(value, (int, float)) and header != "Pipe Size":
                    cell.number_format = '#,##0.00'

    def _create_openpyxl_charts(self, ws_dash, ws_data):
        """Create charts using openpyxl native charts."""
        # Add title
        ws_dash["A1"] = "CIPP Project Dashboard"
        ws_dash["A1"].font = Font(size=20, bold=True)

        ws_dash["A3"] = "Total Segments:"
        ws_dash["B3"] = sum(r["Segment_Count"] for r in self.tables["stage_footage_summary"])

        ws_dash["D3"] = "Total Footage:"
        ws_dash["E3"] = self.processor.total_footage

        # Chart 1: Progress bar (stacked bar)
        chart1 = BarChart()
        chart1.type = "bar"
        chart1.grouping = "percentStacked"
        chart1.overlap = 100
        chart1.title = "Overall Progress"

        data = Reference(ws_data, min_col=3, min_row=1, max_row=7, max_col=3)
        cats = Reference(ws_data, min_col=1, min_row=2, max_row=7)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)

        ws_dash.add_chart(chart1, "A5")

        # Chart 2: Pipe Size (Pie chart)
        chart2 = PieChart()
        chart2.title = "Pipe Size Distribution"

        table3_rows = len(self.tables["pipe_size_mix"]) + 1
        table3_start = 10 + len(self.tables["stage_by_pipe_size"]) + 2

        labels = Reference(ws_data, min_col=1, min_row=table3_start + 1, max_row=table3_start + table3_rows - 1)
        data = Reference(ws_data, min_col=3, min_row=table3_start, max_row=table3_start + table3_rows - 1)

        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(labels)

        ws_dash.add_chart(chart2, "J5")

        # Chart 3: Progress by Pipe Size
        chart3 = BarChart()
        chart3.type = "bar"
        chart3.grouping = "stacked"
        chart3.title = "Progress by Pipe Size"

        table2_rows = len(self.tables["stage_by_pipe_size"]) + 1
        data = Reference(ws_data, min_col=2, min_row=10, max_row=10 + table2_rows - 1, max_col=7)
        cats = Reference(ws_data, min_col=1, min_row=11, max_row=10 + table2_rows - 1)
        chart3.add_data(data, titles_from_data=True)
        chart3.set_categories(cats)

        ws_dash.add_chart(chart3, "A25")

    def _write_all_tables_xlsxwriter(self, wb, ws):
        """Write tables using xlsxwriter."""
        header_format = wb.add_format({
            'bold': True,
            'bg_color': '#4472C4',
            'font_color': 'white',
            'align': 'center'
        })

        number_format = wb.add_format({'num_format': '#,##0.00'})
        percent_format = wb.add_format({'num_format': '0.00%'})

        # Write all tables
        row = 0
        for table_name, table_data in self.tables.items():
            self._write_table_xlsxwriter(ws, row, 0, table_data, header_format, number_format, percent_format)
            row += len(table_data) + 3

    def _write_table_xlsxwriter(self, ws, start_row, start_col, data, header_fmt, num_fmt, pct_fmt):
        """Write a table using xlsxwriter."""
        if not data:
            return

        headers = list(data[0].keys())

        # Write headers
        for col_idx, header in enumerate(headers):
            ws.write(start_row, start_col + col_idx, header, header_fmt)

        # Write data
        for row_idx, row_data in enumerate(data, start=1):
            for col_idx, header in enumerate(headers):
                value = row_data[header]

                if "Pct" in header or "%" in header:
                    ws.write(start_row + row_idx, start_col + col_idx, value, pct_fmt)
                elif isinstance(value, (int, float)) and header != "Pipe Size":
                    ws.write(start_row + row_idx, start_col + col_idx, value, num_fmt)
                else:
                    ws.write(start_row + row_idx, start_col + col_idx, value)

    def _create_xlsxwriter_charts(self, wb, ws_dash, ws_data):
        """Create charts using xlsxwriter."""
        # Add title
        title_format = wb.add_format({'bold': True, 'font_size': 20})
        ws_dash.write('A1', 'CIPP Project Dashboard', title_format)

        # Chart 1: Progress pie
        chart1 = wb.add_chart({'type': 'pie'})
        chart1.add_series({
            'name': 'Progress',
            'categories': '=Dashboard_Data!$A$2:$A$7',
            'values': '=Dashboard_Data!$C$2:$C$7',
            'data_labels': {'percentage': True}
        })
        chart1.set_title({'name': 'Overall Progress'})
        chart1.set_style(10)
        ws_dash.insert_chart('A5', chart1, {'x_scale': 1.5, 'y_scale': 1.5})

    def _create_plotly_charts(self, ws_dash, ws_data):
        """Create plotly charts as images."""
        import plotly.graph_objects as go

        # Add title
        ws_dash["A1"] = "CIPP Project Dashboard"
        ws_dash["A1"].font = Font(size=20, bold=True)

        # Chart 1: Progress donut
        table1 = self.tables["stage_footage_summary"]
        fig1 = go.Figure(data=[go.Pie(
            labels=[r["Stage"] for r in table1],
            values=[r["Total_Feet"] for r in table1],
            hole=0.4,
            marker=dict(colors=['#E74C3C', '#F39C12', '#F1C40F', '#3498DB', '#9B59B6', '#2ECC71'])
        )])
        fig1.update_layout(title="Overall Progress", height=400, width=600)

        img_bytes1 = fig1.to_image(format="png")
        img1 = XLImage(io.BytesIO(img_bytes1))
        ws_dash.add_image(img1, "A5")
